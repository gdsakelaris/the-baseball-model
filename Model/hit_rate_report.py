"""How often do the workbook's stated probabilities come true?

Walks prediction workbook(s), joins every prop cell to the actual box
scores (the exact same PlayerId join grade_results uses), and reports:

  - the headline: props stated OVER 50% -> how many hit vs missed
  - a 10-bucket calibration table (stated probability vs what actually
    happened) across every graded prop cell
  - a per-column breakdown of the over-50% picks

Covers the same surface the grader colors: the 14 batter binary columns,
every pitcher "Stat > line" column, Games "Runs > x", and the Winner /
Win Prob pair. Mean columns (xK, xTB, ...) have no yes/no event and are
skipped. Rows with no box score (scratched player, unscraped finals) are
skipped and counted.

Usage:
    python Model/hit_rate_report.py                    # every workbook in Predictions/
    python Model/hit_rate_report.py path\\to\\file.xlsx [more.xlsx ...]
"""
import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

from grade_results import (BAT_EVENTS, LINE_RE, LINE_STAT, RUNS_RE,
                           PRED_DIR, load_actuals)


def collect_file(path):
    """[(column, stated_p, occurred), ...] for one workbook, plus counts
    of rows skipped for missing box scores."""
    m = re.match(r"(\d{4}-\d{2}-\d{2})", Path(path).stem)
    if not m:
        print(f"  ! {Path(path).name}: no date in filename, skipped")
        return [], 0
    date = m.group(1)
    try:
        batters, starters, games = load_actuals(date)
    except Exception as e:
        print(f"  ! {Path(path).name}: {e}")
        return [], 0
    if not batters:
        print(f"  ! {Path(path).name}: no box scores for {date} yet")
        return [], 0

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows, skipped = [], 0

    def prob(v):
        return v if isinstance(v, (int, float)) and 0 <= v <= 1 else None

    def sheet_head(name):
        ws = wb[name]
        head = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        return ws, {h: j for j, h in enumerate(head)}

    if "Batter Props" in wb.sheetnames:
        ws, hidx = sheet_head("Batter Props")
        if "ID" not in hidx:
            print(f"  ! {Path(path).name}: Batter Props has no ID column "
                  f"(pre-ID workbook), sheet skipped")
            cols = {}
        else:
            cols = {h: j for h, j in hidx.items() if h in BAT_EVENTS}
        for r in (ws.iter_rows(min_row=2, values_only=True) if cols else ()):
            pid = r[hidx["ID"]]
            s = batters.get(int(pid)) if pid is not None else None
            if s is None:
                skipped += 1
                continue
            for h, j in cols.items():
                p = prob(r[j])
                if p is not None:
                    rows.append((h, float(p), bool(BAT_EVENTS[h](s))))

    if "Pitching Props" in wb.sheetnames:
        ws, hidx = sheet_head("Pitching Props")
        line_cols = ([(h, j, LINE_RE.match(h)) for h, j in hidx.items()
                      if LINE_RE.match(h)] if "ID" in hidx else [])
        for r in (ws.iter_rows(min_row=2, values_only=True)
                  if line_cols else ()):
            pid = r[hidx["ID"]]
            a = starters.get(int(pid)) if pid is not None else None
            if a is None:
                skipped += 1
                continue
            for h, j, mm in line_cols:
                p = prob(r[j])
                if p is not None:
                    occ = a[LINE_STAT[mm.group(1)]] > float(mm.group(2))
                    rows.append((h, float(p), bool(occ)))

    if "Games" in wb.sheetnames:
        ws, hidx = sheet_head("Games")
        run_cols = [(h, j, float(RUNS_RE.match(h).group(1)))
                    for h, j in hidx.items() if RUNS_RE.match(h)]
        for r in (ws.iter_rows(min_row=2, values_only=True)
                  if "Game" in hidx else ()):
            g = games.get(str(r[hidx["Game"]]))
            if g is None:
                skipped += 1
                continue
            if "Winner" in hidx and "Win Prob" in hidx:
                p = prob(r[hidx["Win Prob"]])
                if p is not None:
                    rows.append(("Winner", float(p),
                                 str(r[hidx["Winner"]]) == g["winner"]))
            for h, j, line in run_cols:
                p = prob(r[j])
                if p is not None:
                    rows.append((h, float(p), g["total"] > line))
    wb.close()
    return rows, skipped


def report(rows):
    n = len(rows)
    if not n:
        print("nothing to grade.")
        return

    over = [(p, occ) for _, p, occ in rows if p > 0.5]
    under = [(p, occ) for _, p, occ in rows if p <= 0.5]
    o_hit = sum(occ for _, occ in over)
    u_hit = sum(occ for _, occ in under)
    print(f"\n=== Headline: the over-50% picks ===")
    if over:
        avg = sum(p for p, _ in over) / len(over)
        print(f"  stated > 50%:  {len(over):5,} props -> {o_hit:5,} hit "
              f"({o_hit / len(over):6.1%})   stated avg {avg:6.1%}")
    if under:
        avg = sum(p for p, _ in under) / len(under)
        print(f"  stated <= 50%: {len(under):5,} props -> {u_hit:5,} hit "
              f"({u_hit / len(under):6.1%})   stated avg {avg:6.1%}")

    print(f"\n=== Calibration: stated probability vs reality "
          f"({n:,} prop cells) ===")
    print("  stated        props    hit    hit%   stated avg    gap")
    for lo in range(0, 100, 10):
        hi = lo + 10
        b = [(p, occ) for _, p, occ in rows
             if lo / 100 <= p < hi / 100 or (hi == 100 and p == 1.0)]
        if not b:
            continue
        hit = sum(occ for _, occ in b)
        rate, avg = hit / len(b), sum(p for p, _ in b) / len(b)
        print(f"  {lo:3d}-{hi:3d}%   {len(b):7,} {hit:7,} {rate:7.1%}"
              f"    {avg:7.1%}   {rate - avg:+6.1%}")

    print(f"\n=== Over-50% picks by column ===")
    by_col = defaultdict(list)
    for h, p, occ in rows:
        if p > 0.5:
            by_col[h].append((p, occ))
    print("  column           picks    hit    hit%   stated avg    gap")
    for h, b in sorted(by_col.items(), key=lambda kv: -len(kv[1])):
        hit = sum(occ for _, occ in b)
        rate, avg = hit / len(b), sum(p for p, _ in b) / len(b)
        print(f"  {h:14s} {len(b):7,} {hit:7,} {rate:7.1%}"
              f"    {avg:7.1%}   {rate - avg:+6.1%}")


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("workbooks", nargs="*",
                    help=".xlsx paths (default: all of Predictions/)")
    args = ap.parse_args()
    paths = ([Path(p) for p in args.workbooks] if args.workbooks
             else sorted(PRED_DIR.glob("[0-9]*.xlsx")))
    if not paths:
        sys.exit(f"no workbooks found in {PRED_DIR}")

    all_rows, all_skipped = [], 0
    print("workbooks:")
    for p in paths:
        rows, skipped = collect_file(p)
        if rows:
            hits = sum(occ for _, _, occ in rows)
            print(f"  {p.name}: {len(rows):,} prop cells "
                  f"({hits:,} occurred)"
                  + (f", {skipped} row(s) without box scores" if skipped
                     else ""))
        all_rows += rows
        all_skipped += skipped
    report(all_rows)
    if all_skipped:
        print(f"\n{all_skipped} row(s) had no box score (scratched player or "
              f"finals not scraped yet) — not counted.")


if __name__ == "__main__":
    main()
