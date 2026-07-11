"""Grade a predictions workbook against the actual box scores.

Re-colors the workbook IN PLACE once games are final, using the project's
own scraped per-player box scores (Data/mlb_game_batting.csv /
mlb_game_pitching.csv / mlb_games.csv) joined on the workbook's ID column —
PlayerId-exact, so no third-party box-score scraping or name matching:

  stat occurred + cell was plain (white / red tint)  -> yellow
  stat occurred + cell was a light-blue quality pick -> darker blue
  stat occurred + cell was a light-green +EV bet     -> darker green
  stat occurred + cell was light purple (blue + green) -> dark purple

Every graded cell answers the same literal question: DID THE STAT OCCUR.
Binary batter columns light up if the event happened (the HR cell if he
homered); O/U line columns light up if the OVER hit (K > 8.5 only if he
actually struck out 9+); the Winner cell (and its Win Prob) if the named
team won. Mean columns (xK, xTB, Away/Home Score, ...) have no yes/no
event and are left untouched.

Re-running is safe and REPAIRS earlier grades: pass 1 reverts every
occurred-color cell back to its base color (magenta/yellow -> plain, dark blue/
green/purple -> the light pick shade), then pass 2 grades fresh.

Doubleheaders: a batter's stats are summed across the day's games (the
slate carries one row per matchup, so the day total is the honest read).

If some games are missing (they weren't final at the last scrape), their
rows are skipped and counted — run  python Scripts/scrape_gamelogs_3F.py
to pull the late finals, then grade again.

Usage:
    python Model/grade_results.py                    # newest in Predictions/
    python Model/grade_results.py path\\to\\file.xlsx
"""
import argparse
import re
import sys
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill

DATA_DIR = Path(__file__).resolve().parent.parent / "Data"
PRED_DIR = Path(__file__).resolve().parent.parent / "Predictions"

# The visual grammar: LIGHT color = pick pending, DARK fill + white bold =
# pick HIT, pale-gray fill + gray italic = pick MISSED, yellow = a stat
# occurred with no pick on it, white = nothing predicted, nothing happened.
BLUE, GREEN, PURPLE = "00B0F0", "92D050", "B1A0C7"
# earlier palettes, still recognized when repairing old workbooks
OLD_BASES = {"9DC3E6": "00B0F0", "C6E0B4": "92D050",
             "C6EFCE": "92D050", "CCC0DA": "B1A0C7"}
# fill + text color; per-cell font weight follows the bold-over-50% rule
# (a HIT is never bolded unless its stated probability was above 50%)
OCCURRED = {
    BLUE:   (PatternFill("solid", fgColor="0070C0"), "FFFFFF"),  # dark blue
    GREEN:  (PatternFill("solid", fgColor="00B050"), "FFFFFF"),  # dark green
    PURPLE: (PatternFill("solid", fgColor="7030A0"), "FFFFFF"),  # dark purple
    None:   (PatternFill("solid", fgColor="FFFF00"), "000000"),  # yellow
}
# a pick that did NOT hit: grayed out (each fill keeps a faint machine-
# readable hue so re-grading can restore the base color; to the eye they
# all read "gray = miss"). Font is per-cell: probabilities over 50% stay
# bold even in a miss.
MISSED = {
    BLUE:   PatternFill("solid", fgColor="B8CCE4"),
    GREEN:  PatternFill("solid", fgColor="C4D79B"),
    PURPLE: PatternFill("solid", fgColor="D2C8DE"),
}

# batter columns -> did the event happen, from the day's summed line
BAT_EVENTS = {
    "HR":          lambda s: s["HR"] >= 1,
    "Hit":         lambda s: s["H"] >= 1,
    "2+ Hits":     lambda s: s["H"] >= 2,
    "Single":      lambda s: (s["H"] - s["2B"] - s["3B"] - s["HR"]) >= 1,
    "Double":      lambda s: s["2B"] >= 1,
    "2+ TB":       lambda s: s["TB"] >= 2,
    "Run":         lambda s: s["R"] >= 1,
    "RBI":         lambda s: s["RBI"] >= 1,
    "H+R+RBI 2+":  lambda s: (s["H"] + s["R"] + s["RBI"]) >= 2,
    "H+R+RBI 3+":  lambda s: (s["H"] + s["R"] + s["RBI"]) >= 3,
    "BB":          lambda s: s["BB"] >= 1,
    "SB":          lambda s: s["SB"] >= 1,
    "K":           lambda s: s["SO"] >= 1,
    "2+ K":        lambda s: s["SO"] >= 2,
}
BAT_SUM_COLS = ["PA", "AB", "R", "H", "2B", "3B", "HR", "RBI", "BB", "SO",
                "SB", "TB"]

# pitcher O/U column pattern -> the actual-stat key it grades
LINE_RE = re.compile(r"^(K|Outs|Hits|BB|ER) > (\d+(?:\.\d+)?)$")
LINE_STAT = {"K": "SO", "Outs": "outs", "Hits": "H", "BB": "BB", "ER": "ER"}
RUNS_RE = re.compile(r"^Runs > (\d+(?:\.\d+)?)$")


def ip_to_outs(ip):
    """'5.2' -> 17 outs (MLB notation: .1/.2 = thirds)."""
    ip = float(ip)
    whole = int(ip)
    return whole * 3 + round((ip - whole) * 10)


def load_actuals(date):
    """(batters {pid: summed Series}, starters {pid: dict},
    games {'AWY@HOM': dict}) for one date, from the scraped logs."""
    gb = pd.read_csv(DATA_DIR / "mlb_game_batting.csv", encoding="utf-8-sig",
                     low_memory=False)
    gb = gb[gb["Date"] == date]
    batters = {int(pid): grp[BAT_SUM_COLS].sum()
               for pid, grp in gb.groupby("PlayerId")}

    gp = pd.read_csv(DATA_DIR / "mlb_game_pitching.csv", encoding="utf-8-sig",
                     low_memory=False)
    gp = gp[(gp["Date"] == date) & (gp["GS"] == 1)]
    starters = {}
    for pid, grp in gp.groupby("PlayerId"):
        r = grp.iloc[0]
        starters[int(pid)] = {"SO": float(r["SO"]), "H": float(r["H"]),
                              "BB": float(r["BB"]), "ER": float(r["ER"]),
                              "outs": ip_to_outs(r["IP"])}

    g = pd.read_csv(DATA_DIR / "mlb_games.csv", encoding="utf-8-sig")
    g = g[g["Date"] == date].dropna(subset=["AwayScore", "HomeScore"])
    games = {}
    for _, r in g.iterrows():
        a, h = float(r["AwayScore"]), float(r["HomeScore"])
        games[f'{r["AwayTeam"]}@{r["HomeTeam"]}'] = {
            "total": a + h,
            "winner": r["HomeTeam"] if h > a else r["AwayTeam"]}
    return batters, starters, games


# graded-color -> the base to restore on re-grade (yellow was plain).
# Covers tonight's palette AND the earlier one, so previously graded
# workbooks repair cleanly.
UNGRADE = {"0070C0": BLUE, "2E75B6": BLUE,          # hits (new, old)
           "00B050": GREEN, "538135": GREEN, "548235": GREEN,
           "7030A0": PURPLE,
           "FFFF00": None, "CB21C3": None, "FFD966": None, "FFE699": None,
           "B8CCE4": BLUE, "DCE6F1": BLUE, "D8E4BC": GREEN, "C4D79B": GREEN,
           "D2C8DE": PURPLE,
           "E8EDF3": BLUE, "EAF0E4": GREEN, "EDE8F3": PURPLE,
           "9DC3E6": BLUE, "C6E0B4": GREEN,          # old bases
           "C6EFCE": GREEN, "CCC0DA": PURPLE,
           "F5DBE2": None}   # the retired red column tint -> plain white
# base fill + its tinted text color (None = plain cell, default font)
BASE_FILL = {
    BLUE:   (PatternFill("solid", fgColor=BLUE), "0B2E4F"),
    GREEN:  (PatternFill("solid", fgColor=GREEN), "1E4620"),
    PURPLE: (PatternFill("solid", fgColor=PURPLE), "3B2151"),
    None:   (PatternFill(fill_type=None), None),
}


def _ungrade(ws):
    """Revert every occurred-color cell to its base color so grading is
    idempotent (and repairs workbooks graded by the old side-of-the-line
    logic)."""
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            try:
                rgb = cell.fill.start_color.rgb
            except AttributeError:
                continue
            if isinstance(rgb, str) and rgb[-6:].upper() in UNGRADE:
                base = UNGRADE[rgb[-6:].upper()]
                fill, color = BASE_FILL[base]
                cell.fill = fill
                # probabilities follow the bold-over-50% rule everywhere;
                # text cells on a pick fill (e.g. Winner) stay bold
                num = isinstance(cell.value, (int, float))
                bold = (cell.value > 0.5) if num else (base is not None)
                cell.font = (Font(bold=bold, color=color) if color
                             else Font(bold=bold))


def _base_color(cell):
    """The pick color a cell was painted with, or None for plain."""
    try:
        rgb = cell.fill.start_color.rgb
    except AttributeError:
        return None
    if not isinstance(rgb, str):
        return None
    tail = rgb[-6:].upper()
    tail = OLD_BASES.get(tail, tail)
    return tail if tail in (BLUE, GREEN, PURPLE) else None


def _mark(cell, occurred):
    """HIT -> the dark fill; MISS on a pick -> grayed out; MISS on a
    plain cell -> untouched. Font weight follows the bold-over-50% rule
    in EVERY case (hit or miss); text cells (e.g. Winner) stay bold on
    fills for contrast."""
    base = _base_color(cell)
    num = isinstance(cell.value, (int, float))
    bold = bool(cell.value > 0.5) if num else True
    if occurred:
        fill, color = OCCURRED[base]
        cell.fill = fill
        cell.font = Font(bold=bold, color=color)
        return
    if base is None:
        return
    cell.fill = MISSED[base]
    cell.font = Font(italic=True, color="7F7F7F", bold=num and bold)




def grade(path):
    m = re.match(r"(\d{4}-\d{2}-\d{2})", Path(path).stem)
    if not m:
        sys.exit(f"can't read the game date from the filename: {path}")
    date = m.group(1)
    batters, starters, games = load_actuals(date)
    if not batters:
        sys.exit(f"no box scores for {date} in Data/mlb_game_batting.csv — "
                 f"run  python Scripts/scrape_gamelogs_3F.py  first")

    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        _ungrade(ws)
    stats = {"cells": 0, "hit": 0, "missing_rows": 0}

    def headers_of(ws):
        return {str(c.value): j for j, c in enumerate(ws[1], start=1)}

    if "Batter Props" in wb.sheetnames:
        ws = wb["Batter Props"]
        hidx = headers_of(ws)
        cols = {h: j for h, j in hidx.items() if h in BAT_EVENTS}
        for i in range(2, ws.max_row + 1):
            pid = ws.cell(row=i, column=hidx["ID"]).value
            s = batters.get(int(pid)) if pid is not None else None
            if s is None:
                stats["missing_rows"] += 1
                continue
            for h, j in cols.items():
                stats["cells"] += 1
                occ = bool(BAT_EVENTS[h](s))
                stats["hit"] += occ
                _mark(ws.cell(row=i, column=j), occ)

    if "Pitching Props" in wb.sheetnames:
        ws = wb["Pitching Props"]
        hidx = headers_of(ws)
        line_cols = [(h, j, LINE_RE.match(h)) for h, j in hidx.items()
                     if LINE_RE.match(h)]
        for i in range(2, ws.max_row + 1):
            pid = ws.cell(row=i, column=hidx["ID"]).value
            a = starters.get(int(pid)) if pid is not None else None
            if a is None:
                stats["missing_rows"] += 1
                continue
            for h, j, mm in line_cols:
                stats["cells"] += 1
                occ = a[LINE_STAT[mm.group(1)]] > float(mm.group(2))
                stats["hit"] += occ
                _mark(ws.cell(row=i, column=j), occ)

    if "Games" in wb.sheetnames:
        ws = wb["Games"]
        hidx = headers_of(ws)
        run_cols = [(j, float(RUNS_RE.match(h).group(1)))
                    for h, j in hidx.items() if RUNS_RE.match(h)]
        for i in range(2, ws.max_row + 1):
            tag = ws.cell(row=i, column=hidx["Game"]).value
            g = games.get(str(tag))
            if g is None:
                stats["missing_rows"] += 1
                continue
            if "Winner" in hidx:
                cell = ws.cell(row=i, column=hidx["Winner"])
                stats["cells"] += 1
                occ = str(cell.value) == g["winner"]
                stats["hit"] += occ
                _mark(cell, occ)
                # Win Prob belongs to the named winner -> same outcome
                if "Win Prob" in hidx:
                    _mark(ws.cell(row=i, column=hidx["Win Prob"]), occ)
            for j, line in run_cols:
                stats["cells"] += 1
                occ = g["total"] > line
                stats["hit"] += occ
                _mark(ws.cell(row=i, column=j), occ)

    try:
        wb.save(path)
    except PermissionError:
        sys.exit(f"{path} is open in Excel (it holds the file lock) — "
                 f"close it there, then run this again.")
    return date, stats


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("workbook", nargs="?", default=None,
                    help="predictions .xlsx (default: newest in Predictions/)")
    args = ap.parse_args()
    path = args.workbook
    if path is None:
        books = sorted(PRED_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime)
        if not books:
            sys.exit(f"no workbooks in {PRED_DIR}")
        path = books[-1]
    date, s = grade(path)
    print(f"graded {path}")
    print(f"  {date}: {s['cells']:,} cells checked, {s['hit']:,} stats "
          f"occurred (now yellow / dark blue / dark green / dark purple)")
    if s["missing_rows"]:
        print(f"  {s['missing_rows']} row(s) had no final box score yet — "
              f"run  python Scripts/scrape_gamelogs_3F.py  and grade again")


if __name__ == "__main__":
    main()
