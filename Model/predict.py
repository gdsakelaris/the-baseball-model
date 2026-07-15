"""Prediction engine: turn a game spec into betting-relevant predictions.

Given a game (teams, date, venue, starters, ordered lineups, weather), the
Predictor outputs, per run:
  - every lineup batter's calibrated P(home run), P(1+ hit), expected HRs,
    and fair American odds for the HR prop
  - each starter's expected strikeouts and P(over) for common K lines
  - game totals: expected lineup home runs and expected total runs

Usage:
    python Model/predict.py --game 745444      # replay a historical game
    python Model/predict.py --selftest         # train/serve parity check
"""

import argparse
import math
import re
import sys
from collections import defaultdict
from pathlib import Path

import joblib
import numpy as np
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))
import features as F  # noqa: E402
import odds as O  # noqa: E402
import recalibrate as R  # noqa: E402

ART = Path(__file__).resolve().parent / "artifacts"

K_LINES = [3.5, 4.5, 5.5, 6.5, 7.5, 8.5]
TOTAL_LINES = [6.5, 7.5, 8.5, 9.5, 10.5]
# per-TEAM total-runs lines (H5 team_total head, 2026-07-14)
TEAM_TOTAL_LINES = [2.5, 3.5, 4.5, 5.5]

# prop key -> output column, in display order (2026-07-14: + the H1 deep
# binaries, H3 triple, H4 2+ RBI / 2+ runs)
PROP_COLS = {"hr": "P_HR", "hit": "P_Hit", "hits2": "P_2Hits",
             "tb2": "P_TB2", "tb3": "P_TB3", "tb4": "P_TB4",
             "run": "P_Run", "run2": "P_Run2",
             "rbi": "P_RBI", "rbi2": "P_RBI2",
             "bb": "P_BB", "sb": "P_SB",
             "single": "P_1B", "double": "P_2B", "triple": "P_3B",
             "bk": "P_K", "bk2": "P_2K", "bk3": "P_3K",
             "hrr2": "P_HRR2", "hrr3": "P_HRR3", "hrr4": "P_HRR4"}

# batter count heads -> mean column; starter count heads -> (mean, P prefix)
# (2026-07-14: + the H6 expected-stat-line means — means ONLY, their line
# calibrators stay banked; binaries own the batter lines)
BAT_COUNT_COLS = {"xbk": "xSO", "xhrr": "xHRR", "xtb": "xTB",
                  "xh": "xH", "xrun": "xR", "xrbi": "xRBI", "xbb": "xBB"}
ST_COUNT_COLS = {"outs": ("xOuts", "P_outs_over"),
                 "pbb": ("xBB", "P_bb_over"),
                 "pha": ("xHits", "P_hits_over"),
                 "per": ("xER", "P_er_over")}


def american_odds(p):
    if not (0 < p < 1):
        return ""
    if p >= 0.5:
        return f"-{round(100 * p / (1 - p)):d}"
    return f"+{round(100 * (1 - p) / p):d}"


def poisson_over(lam, line):
    """P(count > line) for a half-point line under Poisson(lam)."""
    k = int(math.floor(line))
    cdf = sum(math.exp(-lam) * lam ** i / math.factorial(i) for i in range(k + 1))
    return 1 - cdf


def nb_over(lam, line, disp=1.0):
    """P(count > line) when variance = disp * mean (negative binomial).

    Real MLB game totals are ~2x more variable than Poisson (blowouts,
    extras, shared park/weather), which made pure-Poisson P(over) worse
    than the base rate at low run lines. `disp` is measured on the
    calibration year at train time. Falls back to Poisson at disp <= 1
    (starter strikeouts measured ~1.04, so K lines stay Poisson)."""
    if disp <= 1.001 or lam <= 0:
        return poisson_over(lam, line)
    r = lam / (disp - 1.0)          # NB: mean lam, variance disp*lam
    log_p, log_q = -math.log(disp), math.log(1.0 - 1.0 / disp)
    lg_r = math.lgamma(r)
    k = int(math.floor(line))
    cdf = sum(math.exp(math.lgamma(i + r) - lg_r - math.lgamma(i + 1)
                       + r * log_p + i * log_q) for i in range(k + 1))
    return 1 - cdf


# Heads whose lines are priced by the negative binomial even though the
# artifact carries per-line calibrators. per (earned runs, disp ~1.5) moved
# here after the 2026-07-09 shoot-out: NB beat the calibrators on BOTH tail
# lines (3.5/4.5) in BOTH years and on every 2026 line (mean rel-edge
# +0.11), tying 2025 overall — and a single distribution can't produce
# crossing line probabilities the way independent per-line logistics can.
# Under-dispersed heads (outs ~0.8) stay on calibrators: NB can only widen
# variance. One-line revert: remove the target from this set.
# 2026-07-13 full-surface pricing comparison (count_pricing_compare, all 9
# families x both years, day-block CIs): per's NB verdict RE-CONFIRMED on
# the ensemble-era model (2026 NB CI-clear at 2.5/3.5, 2025 flat); K added
# (NB CI-clear at the low lines both years, calibrators win nowhere — K is
# ~Poisson, disp 1.01); total added (calibrator's lone 2025 win at 7.5
# flipped sign on 2026 + a CI-clear 2026 NB win at 6.5). Their fitted
# calibrators stay banked in the artifacts (k_line_cals/total_line_cals);
# removing a target from this set adopts them.
NB_PRICED_TARGETS = {"y_per", "y_so", "total_runs"}

# PA-sim game-level blend (Phase 3, 2026-07-13): w_sim per head, applied
# to the sim outputs from Model/pa_serve.py — score/total means linear,
# winner on logits. Empty dict (or missing sim artifacts) = incumbent
# alone; batter/starter heads and sb stay incumbent (w=0 verdicts —
# reaffirmed 07-15 with the #35 battery live: sb fit-2025 w still 0).
# Weights = the 2026-07-15 finish-chain pa_blend fit on 2025 ONLY
# (hazard v2 engine, battery live), 2026 confirm-only read showed no
# CI-clear harm; USER decision 07-15 = wire the fit-2025 verbatim. This
# re-decision RESOLVES the audit-#7 ledgered exception on total (its old
# 0.20 was a 2026-informed half-weight tune; the fresh 2025-only fit
# says 0.55). NOTE: the first table shown carried a stale-parquet
# duplication (pa_grade._sim glob matched the superseded 07-14 combined
# files alongside the 07-15 parts — fixed in pa_grade); the clean re-fit
# moved only score (0.65 -> 0.60), and the ratified verbatim rule was
# applied to the corrected fit.
SIM_BLEND = {"score": 0.60, "total": 0.55, "winner": 0.30}


def _sim_logit(p):
    p = min(max(float(p), 1e-6), 1 - 1e-6)
    return np.log(p / (1 - p))


def count_over(head, mu, line):
    """P(count > line) for a count head: the calibration-year per-line
    logistic when the artifact carries one (under-dispersed counts like
    starter outs misprice under NB, which can only widen variance), else
    nb_over with the head's dispersion. Over-dispersed heads listed in
    NB_PRICED_TARGETS skip their calibrators and price NB (fat tails).
    Returns an array over mu."""
    mu = np.atleast_1d(np.asarray(mu, dtype=float))
    if head.get("target") in NB_PRICED_TARGETS:
        return np.array([nb_over(m, line, head["disp"]) for m in mu])
    lc = head.get("line_cals", {}).get(line)
    if lc is not None:
        return lc.predict_proba(mu.reshape(-1, 1))[:, 1]
    return np.array([nb_over(m, line, head["disp"]) for m in mu])


def k_over(art, mu, line):
    """P(starter K > line): the cal-year per-line logistic when the artifact
    carries one (2026-07-13 full-surface calibration pass — K lines were the
    only starter family still priced by the raw NB tail), else nb_over with
    the cal-year K dispersion (old-artifact guard)."""
    mu = np.atleast_1d(np.asarray(mu, dtype=float))
    if len(mu) == 0:
        return np.array([])
    lc = (None if "y_so" in NB_PRICED_TARGETS
          else art.get("k_line_cals", {}).get(line))
    if lc is not None:
        return lc.predict_proba(mu.reshape(-1, 1))[:, 1]
    disp = float(art.get("k_disp", 1.0))
    return np.array([nb_over(m, line, disp) for m in mu])


def _vmr_scaled_disp(art, disp, park_vmr):
    """Audit wave rank 32: scale the NB total dispersion by the venue's
    run variance-to-mean ratio, disp * (park_vmr/VMR0)^a. FAIL-SAFE and
    GATED: the exponent `a` is fit on 2025 ONLY (evaluate_deep, Phase-5)
    and stored as art['total_vmr_exp']; when it is absent (default) or
    park_vmr is missing, this returns the constant disp unchanged, so the
    column ships live but the pricing change only activates once the 2025
    fit lands and clears the paired gate."""
    a = art.get("total_vmr_exp")
    if a is None or park_vmr is None or not np.isfinite(park_vmr):
        return disp
    from features import PARK_VMR0
    scaled = disp * (float(park_vmr) / PARK_VMR0) ** float(a)
    return float(np.clip(scaled, 1.6, 3.0))


def total_over(art, mu, line, park_vmr=None):
    """P(game total runs > line): cal-year per-line logistic when present
    (same 2026-07-13 pass — total lines were the worst-calibrated family on
    the board under the raw NB tail), else nb_over with the cal-year total
    dispersion (venue-scaled when the gated 2025 vmr fit is present). Lines
    outside TOTAL_LINES (odds-store exotics) always fall back to NB."""
    mu = np.atleast_1d(np.asarray(mu, dtype=float))
    if len(mu) == 0:
        return np.array([])
    lc = (None if "total_runs" in NB_PRICED_TARGETS
          else art.get("total_line_cals", {}).get(line))
    if lc is not None:
        return lc.predict_proba(mu.reshape(-1, 1))[:, 1]
    disp = _vmr_scaled_disp(art, float(art.get("total_disp", 1.0)), park_vmr)
    return np.array([nb_over(m, line, disp) for m in mu])


def team_over(art, mu, line):
    """P(one TEAM's runs > line) — the H5 team_total head's line surface
    (2026-07-14): cal-year per-line logistic when the artifact carries one,
    else nb_over with the TEAM-level dispersion (the game total's does not
    transfer). Old artifacts (no team cals/disp) degrade to Poisson."""
    mu = np.atleast_1d(np.asarray(mu, dtype=float))
    if len(mu) == 0:
        return np.array([])
    lc = art.get("team_line_cals", {}).get(line)
    if lc is not None:
        return lc.predict_proba(mu.reshape(-1, 1))[:, 1]
    disp = float(art.get("team_total_disp", 1.0))
    return np.array([nb_over(m, line, disp) for m in mu])


def apply_stack(prop, p_self, donor_ps):
    """Calibration-layer stacking for thin-signal props: a logistic (fit on
    the calibration year by train.py, never the holdout) blends the prop's
    own calibrated probability with its donor props' scores in log-odds
    space. donor_ps maps prop name -> probability array for the SAME rows.
    Props without a "stack" key, or with a donor missing from donor_ps,
    pass through unchanged (old-artifact guard)."""
    st = prop.get("stack") if isinstance(prop, dict) else None
    if not st or any(d not in donor_ps for d in st["donors"]):
        return p_self
    Z = np.column_stack([R._logit(p_self)]
                        + [R._logit(donor_ps[d]) for d in st["donors"]])
    return np.clip(st["lr"].predict_proba(Z)[:, 1], 1e-4, 1 - 1e-4)


def predict_prop(prop, X):
    """Calibrated probability from a prop's GBM+logistic blend. X is prepped
    with the full batter column set; props trained on a subset (per-prop
    feature selection) carry their own column list. Backwards-compatible
    with older {model, iso} artifacts. blend_space (2026-07-15) says which
    space the blend weight was fit in — log-odds for new artifacts; absent
    key = old artifact = probability space."""
    if "gbm" in prop:
        Xp = X[prop["cols"]] if "cols" in prop else X
        g = prop["gbm"].predict_proba(Xp)[:, 1]
        l = prop["lr"].predict_proba(Xp[prop["lr_cols"]])[:, 1]
        s = F.blend(g, l, prop["w"], prop.get("blend_space", "prob"))
        return prop["iso"].predict(s)
    return prop["iso"].predict(prop["model"].predict_proba(X)[:, 1])


def poisson_win(mu_a, mu_b, kmax=30):
    """P(team A outscores team B) for independent Poisson scores; ties
    (games headed to extras) are split evenly."""
    pa = [math.exp(-mu_a) * mu_a ** i / math.factorial(i) for i in range(kmax)]
    pb = [math.exp(-mu_b) * mu_b ** i / math.factorial(i) for i in range(kmax)]
    win = sum(pa[i] * sum(pb[:i]) for i in range(1, kmax))
    tie = sum(pa[i] * pb[i] for i in range(kmax))
    return win + 0.5 * tie


def predict_win(win_art, X, mu_home, mu_away):
    """Home-win probability from the winner artifact: GBM+logistic blend,
    then (when trained with one) a second blend with the runs-model Poisson
    win probability, then isotonic calibration. Backwards-compatible with
    artifacts that have no Poisson blend (w_ml defaults to 1) and with
    pre-blend_space artifacts (absent key = probability space)."""
    space = win_art.get("blend_space", "prob")
    g = win_art["gbm"].predict_proba(X[win_art["cols"]])[:, 1]
    l = win_art["lr"].predict_proba(X[win_art["lr_cols"]])[:, 1]
    s = F.blend(g, l, win_art["w"], space)
    w_ml = win_art.get("w_ml", 1.0)
    if w_ml < 1.0:
        pois = np.array([poisson_win(h, a) for h, a in
                         zip(np.atleast_1d(mu_home), np.atleast_1d(mu_away))])
        s = F.blend(s, np.where(np.isfinite(pois), pois, s), w_ml, space)
    return win_art["iso"].predict(s)


def _force_xgb_cpu(obj, _seen=None):
    """Move every XGBoost bag member to CPU for INFERENCE. train.py fits XGB
    with device='cuda', and the sklearn wrapper carries that into serving, so
    predict_proba runs on the GPU — which hard-crashes the whole process (taking
    the GUI window with it) when the GPU is busy with a training job or absent.
    Slate inference is a few hundred rows: trivially fast on CPU. This walks the
    loaded art (MeanBag.models + dict/list/tuple containers) and flips each XGB
    booster's device. Serving-only: evaluate_deep scores models directly, never
    through Predictor, so trained baselines are untouched. Returns the count
    flipped. CatBoost is skipped — it applies models on CPU regardless."""
    if _seen is None:
        _seen = set()
    if id(obj) in _seen:
        return 0
    _seen.add(id(obj))

    def _flip(m):
        try:
            m.set_params(device="cpu")
            m.get_booster().set_param({"device": "cpu"})
            return 1
        except Exception:
            return 0

    # InfSafe-wrapped (train.py always wraps XGB) or a bare XGB model
    inner = getattr(obj, "model", None)
    if inner is not None and type(inner).__module__.split(".")[0] == "xgboost":
        return _flip(inner)
    if type(obj).__module__.split(".")[0] == "xgboost":
        return _flip(obj)

    n = 0
    members = getattr(obj, "models", None)   # MeanBag
    if isinstance(members, (list, tuple)):
        for m in members:
            n += _force_xgb_cpu(m, _seen)
    if isinstance(obj, dict):
        for v in obj.values():
            n += _force_xgb_cpu(v, _seen)
    elif isinstance(obj, (list, tuple)):
        for v in obj:
            n += _force_xgb_cpu(v, _seen)
    return n


class Predictor:
    def __init__(self, stores=None, progress=None, recal=False):
        tick = progress or (lambda msg: None)
        tick("loading models...")
        self.art = joblib.load(ART / "models.joblib")
        # serving guard (audit #8, approved 07-15): a superset/experiment
        # intermediate must never serve. Refuse on the role stamp or on any
        # shdw_ column in a serving contract, with a clear message instead
        # of silently NaN-imputing shadow features.
        stamp = self.art.get("meta_stamp") or {}
        serving_lists = {"bat_cols": self.art.get("bat_cols", []),
                         "st_cols": self.art.get("st_cols", []),
                         "tg_cols": self.art.get("tg_cols", [])}
        for n, p in (self.art.get("props") or {}).items():
            serving_lists[f"prop:{n}"] = p.get("cols", [])
        shadowed = sorted(n for n, cols in serving_lists.items()
                          if any(str(c).startswith("shdw_") for c in cols))
        if stamp.get("role") == "superset" or shadowed:
            raise SystemExit(
                f"REFUSING to serve models.joblib: it is a SUPERSET/"
                f"experiment intermediate (role={stamp.get('role')!r}, "
                f"shadowed contracts: {shadowed[:5]}"
                f"{'...' if len(shadowed) > 5 else ''}). Run the keep-train "
                f"(train.py with feature_keep.json present) to restore a "
                f"servable artifact. [audit #8 guard]")
        # serving XGB inference on CPU -> immune to GPU contention/absence
        _force_xgb_cpu(self.art)
        # opt-in in-season drift correction (evaluate_deep Section 10 is the
        # evidence for turning it on). 2026-07-15 (audit #16): offsets live
        # in a sidecar JSON, not inside the serving artifact; the art key is
        # only the pre-07-15 fallback.
        self.recal = recal
        self.offsets = {}
        off_path = ART / "inseason_offsets.json"
        if off_path.exists():
            try:
                import json as _json
                self.offsets = _json.loads(
                    off_path.read_text()).get("offsets", {})
            except Exception:
                self.offsets = {}
        if not self.offsets:
            self.offsets = self.art.get("inseason_offsets") or {}
        self.stores = stores or F.Stores(progress=progress)
        self._slate_sim = None      # lazy PA-sim (False = unavailable)
        tick("indexing names...")
        # Name lookup, broadest to most current: batting AND pitching logs
        # (pitchers never bat post-DH, so batting logs alone miss them),
        # then the handedness file, then rosters.
        self.names = {}
        for frame in (self.stores.raw["gb"], self.stores.raw["gp"]):
            last = frame.sort_values("Date").groupby("PlayerId")["Name"].last()
            self.names.update(last.to_dict())
        hands = self.stores.raw["hands"]
        for pid, name in zip(hands["PlayerId"], hands["Name"]):
            if name:
                self.names.setdefault(pid, name)
        for pid, row in self.stores.rosters.iterrows():
            self.names[pid] = row["Name"]

    def _name(self, pid, spec=None):
        """Display name; falls back to names scraped into the game spec
        (covers debut players with no MLB history yet), then the raw id."""
        n = self.names.get(pid)
        if n:
            return n
        names = (spec or {}).get("names") or {}
        return names.get(str(pid)) or names.get(pid) or str(pid)

    # ---------------------------------------------------------- rows

    def _weather(self, spec):
        """Missing weather inputs become NaN/blank; the models handle both."""
        def num(v):
            try:
                return float(v)
            except (TypeError, ValueError):
                return np.nan
        def txt(v):
            return str(v).strip().title() if v not in (None, "") else ""
        return {"Temp": num(spec.get("temp")),
                "WindSpeed": num(spec.get("wind_speed")),
                "WindDir": txt(spec.get("wind_dir")),
                "Condition": txt(spec.get("condition")),
                "Humidity": num(spec.get("humidity")),
                "Pressure": num(spec.get("pressure")),
                "Precip": num(spec.get("precip")),
                "DayNight": spec.get("day_night") or ""}

    def _sched(self, spec, team):
        """Schedule context for one team (2026-07-14 #18/#24): previous-game
        travel/day-night from the games history + tonight's doubleheader
        flags from the spec (the slate scrape knows the game number;
        historical replays derive it in spec_from_game)."""
        out = self.stores.team_sched(team, pd.Timestamp(spec["date"]),
                                     spec["venue"], spec.get("day_night"))
        out["is_dh"] = float(spec.get("is_dh") or 0.0)
        out["dh_game2"] = float(spec.get("dh_game2") or 0.0)
        return out

    def _batter_rows(self, spec):
        date = pd.Timestamp(spec["date"])
        season = date.year
        s = self.stores
        wx = self._weather(spec)
        park = s.park(spec["venue"], date)
        env = s.league_env(date)
        # audit wave ranks 4/24: prior-season league SB/27 (regime centering
        # for the steal-permissiveness products), game-level
        lg_sb27 = {"lg_sb27_prior": s.lg_sb27_prior(season)}

        # HP-umpire tendency: one game-level value shared by every batter
        ump = s.ump_feats(spec.get("hp_ump_id"), date)

        # the ACTUAL defense each side bats against: player-level prior-season
        # OAA of the OTHER side's posted lineup (mirrors _lineup_oaa_table)
        ldef = {"away": s.lineup_oaa([p for p, _ in spec["home_lineup"]],
                                     season),
                "home": s.lineup_oaa([p for p, _ in spec["away_lineup"]],
                                     season)}

        rows, meta = [], []
        sides = [("away", spec["away_team"], spec["home_team"],
                  spec["away_lineup"], spec["home_starter"], 0),
                 ("home", spec["home_team"], spec["away_team"],
                  spec["home_lineup"], spec["away_starter"], 1)]
        for side, team, opp, lineup, opp_starter, home in sides:
            # unknown opposing starter -> sentinel id; every starter-derived
            # feature (incl. arsenal matchup, platoon) comes back NaN
            opp_starter = opp_starter if opp_starter else -1
            # the opposing starter's venue context is the mirror of ours
            st_feats = s.starter_feats(opp_starter, date, season,
                                       home=1 - home)
            sched = self._sched(spec, team)
            toff = s.team_offense(team, season, date)
            toff_loc = s.team_offense_loc(team, season, home, date)
            pen = s.bullpen(opp, season, date)
            pen_hl = s.bullpen_hl(opp, season, date)
            pen_fat = {"pen_np_l3": s.pen_fatigue(opp, date)}
            tsb = s.team_sb_allowed(opp, season, date)
            phrq = s.pitcher_hr_quality(opp_starter, date)
            pbip = s.bip_pitcher(opp_starter, date)
            pdo = s.pd_pitcher_feats(opp_starter, date)
            tto = s.tto(opp_starter, date)
            oaa_opp = s.team_oaa(opp, season)
            uer_opp = s.team_defense(opp, season, date)
            pprior = s.pitcher_prior(opp_starter, season)
            p_bio = s.bio(opp_starter)
            opp_hand = p_bio["pit_throws"]
            # audit wave: the opposing starter's MiLB prior (rank 17) +
            # arsenal breadth/trajectory (#22), this team's BaseRuns luck
            # (rank 15) + the opponent's HL-arm availability (#22)
            pmilb_opp = s.milb_feats(opp_starter, season, "pit")
            arsdyn_opp = s.arsenal_dynamics(opp_starter, season)
            side_extra = {"toff_bsr_luck": s.team_bsr(team, date),
                          "pen_hl_unavail": s.pen_unavail(opp, date),
                          "p_age": ((date - p_bio["dob"]).days / 365.25
                                    if pd.notna(p_bio["dob"]) else np.nan),
                          "p_effprem_d": pdo["pd_effprem_d"],
                          "p_brkmov_d": pdo["pd_brkmov_d"],
                          "p_stretch_d": pdo["pd_stretch_vdelta_d"],
                          "p_relsep_d": pdo["pd_relsep_d"],
                          # battery + IL wave (2026-07-15): the opponent's
                          # battery quality + the opposing starter's IL
                          # return context (side-level, same for all nine)
                          **s.catcher_feats(opp, season, "opp"),
                          **s.il_feats(opp_starter, date, "p_")}
            side_rows = []
            for pid, slot in lineup:
                b = s.batter_feats(pid, date, season, opp_hand=opp_hand,
                                   home=home)
                bio = s.bio(pid)
                row = {"slot": slot, "Home": home, "Season": season,
                       "month": date.month,
                       "xpa_slot": s.xpa_slot(slot, date), **ump, **sched,
                       **b, **st_feats, **toff,
                       **toff_loc, **pen, **pen_hl, **pen_fat, **tsb, **phrq,
                       **pbip, **s.bip_batter(pid, date),
                       **s.pd_batter_feats(pid, date),
                       "p_swstr_d": pdo["pd_swstr_d"],
                       "p_fbv_d": pdo["pd_fbv_d"],
                       "p_zone_d": pdo["pd_zone_d"],
                       "p_chase_d": pdo["pd_chase_d"],
                       "p_brk_d": pdo["pd_brk_d"],
                       "p_off_d": pdo["pd_off_d"],
                       "p_edge_d": pdo["pd_edge_d"],
                       "p_fps_d": pdo["pd_fps_d"],
                       "p_fblou_d": pdo["pd_fblou_d"],
                       "p_fbmidu_d": pdo["pd_fbmidu_d"],
                       "p_fb95u_d": pdo["pd_fb95u_d"],
                       "p_tswh_d": pdo["pd_tswh_d"],
                       "p_f32b_d": pdo["pd_f32b_d"],
                       "p_fbv_sd": pdo["pd_fbv_sd"],
                       "p_rel_sd": pdo["pd_rel_sd"],
                       "p_ivb_d": pdo["pd_ivb_d"],
                       # v8: his damage-allowed splits (band + class)
                       "p_fbloxw_d": pdo["pd_fbloxw_d"],
                       "p_fbmidxw_d": pdo["pd_fbmidxw_d"],
                       "p_fb95xw_d": pdo["pd_fb95xw_d"],
                       "p_brkxw_d": pdo["pd_brkxw_d"],
                       "p_offxw_d": pdo["pd_offxw_d"],
                       "p_fbkxw_d": pdo["pd_fbkxw_d"],
                       "opp_oaa": oaa_opp, "opp_def_uer": uer_opp,
                       **lg_sb27, **pmilb_opp, **arsdyn_opp, **side_extra,
                       **ldef[side], **tto,
                       **s.batter_brr(pid, season),
                       **s.il_feats(pid, date),
                       **pprior, **park, **wx, **env,
                       "hrpt_score": s.hrpt(pid, opp_starter, season, date),
                       **s.fatigue(pid, date),
                       **s.bvp(pid, opp_starter, date),
                       **s.park_hand_hr(spec["venue"], date),
                       "bat_hand": bio["bat_hand"], "pit_throws": p_bio["pit_throws"],
                       "bat_height": bio["height"], "bat_weight": bio["weight"],
                       "bat_age": ((date - bio["dob"]).days / 365.25
                                   if pd.notna(bio["dob"]) else np.nan)}
                bh, pt = bio["bat_hand"], p_bio["pit_throws"]
                eff = np.nan
                if pd.isna(bh):
                    row["same_hand"] = np.nan
                elif bh == "S":
                    eff = ("R" if pt == "L" else "L") if pd.notna(pt) else np.nan
                    row["same_hand"] = float(eff == pt) if pd.notna(pt) else np.nan
                else:
                    eff = bh
                    row["same_hand"] = float(eff == pt) if pd.notna(pt) else np.nan
                # eff_hand feeds add_batter_derived (called on the assembled
                # frame in predict_game), which computes pull_fence/porch_margin,
                # park_hand_hr_edge, wind carry, temp x elevation, bip x defense,
                # hit_luck, the ump interactions and the BvP residuals — the same
                # shared function the training frame uses, so serving matches it.
                row["eff_hand"] = eff
                # hand-split contact quality (needs opp_hand / eff hand)
                row.update(s.bip_batter_vs_hand(pid, date, opp_hand))
                row.update(s.bip_pitcher_vs_hand(opp_starter, date, eff))
                # form/trend deltas (same helpers as the training frames)
                F.add_bat_trends(row)
                F.add_pit_trends(row)
                side_rows.append(row)
                meta.append({"Team": team, "PlayerId": pid, "slot": slot,
                             "Name": self._name(pid, spec),
                             "BatterId": pid, "PitcherId": opp_starter,
                             "Season": season})
            # teammate context: career + decayed on-base of the two slots
            # ahead, slugging of the two behind, wrapping the order (mirrors
            # the vectorized ctx_* computation; missing slots skipna)
            ctx_maps = {"ctx_ahead_obp": ("_obpp", (-2, -1)),
                        "ctx_behind_slg": ("_slgp", (1, 2)),
                        "ctx_ahead_obp_d": ("_obpp_d", (-2, -1)),
                        "ctx_behind_slg_d": ("_slgp_d", (1, 2)),
                        # runners-ahead advancement (RBI): neighbors'
                        # prior-season extra-base rate, from batter_brr
                        "ctx_ahead_brr": ("bat_brr_xb", (-2, -1)),
                        # audit wave rank 12: the missing neighbor directions
                        # — ahead-SLG (which base the runners occupy),
                        # behind-OBP (chain continuation), behind-GIDP (the
                        # rally-kill axis)
                        "ctx_ahead_slg": ("_slgp", (-2, -1)),
                        "ctx_ahead_slg_d": ("_slgp_d", (-2, -1)),
                        "ctx_behind_obp": ("_obpp", (1, 2)),
                        "ctx_behind_obp_d": ("_obpp_d", (1, 2)),
                        "ctx_behind_gidp": ("c_gidp_pa_sh", (1, 2))}

            def _nmean(vals):
                vals = [v for v in vals if v is not None and pd.notna(v)]
                return float(np.mean(vals)) if vals else np.nan

            for feat, (src, offs) in ctx_maps.items():
                vmap = {r["slot"]: r.get(src) for r in side_rows}
                for r in side_rows:
                    r[feat] = _nmean(
                        [vmap.get(((r["slot"] + off - 1) % 9) + 1)
                         for off in offs])
            # rbi opportunity: full-order decayed OBP of the hitters ahead
            # (mirrors the vectorized rbi_opp_obp; features.RBI_OPP_AHEAD)
            obp_by_slot = {r["slot"]: r.get("_obpp") for r in side_rows}
            for r in side_rows:
                r["rbi_opp_obp"] = F.rbi_opp_from_slots(obp_by_slot, r["slot"])
            rows.extend(side_rows)
        df = pd.DataFrame(rows)
        mdf = pd.DataFrame(meta)
        # arsenal matchup features for the 18 pairs
        mu = F.matchup_features(mdf[["BatterId", "PitcherId", "Season"]],
                                self.stores.raw["ars_p"], self.stores.raw["ars_b"])
        mdf2 = mdf.merge(mu, on=["BatterId", "PitcherId", "Season"], how="left")
        for c in [*F.ARS_P_METRICS.values(), *F.ARS_B_METRICS.values(), "m_coverage"]:
            df[c] = mdf2[c].to_numpy()
        # row-wise derived features + interactions, from the SAME shared function
        # the training frame uses (parity) — pull_fence/porch_margin, wind carry,
        # carry_air, bip x defense, hit_luck, ump interactions, park_hand_hr_edge,
        # BvP residuals — off the as-of inputs assembled above. In _batter_rows so
        # both predict_game and the selftest see them.
        df = F.add_batter_derived(df)
        return df, mdf

    _LU_COLS = {"lu_k_sh": "s_k_pct_sh", "lu_bb_sh": "s_bb_pct_sh",
                "lu_whiff": "m_whiff", "lu_vsh_k": "vsh_k_pct_sh",
                "lu_wsw": "bd_wsw_d", "lu_chase": "bd_chase_d",
                "lu_brkwh": "bd_brkwh_d", "lu_offwh": "bd_offwh_d",
                "lu_fbwh": "bd_fbwh_d",
                # audit wave: rank 18 (damage/OBP/first-pitch), rank 5c
                # (rally-kill), rank 24 (running game) — all means of
                # already-served batter columns
                "lu_obp": "_obpp", "lu_slg": "_slgp",
                "lu_xwcon": "bipd_xwoba", "lu_brl": "bipd_brl",
                "lu_fpsw": "bd_fpsw_d", "lu_gidp": "c_gidp_pa_sh",
                "lu_sb": "d_sb_pa_sh", "lu_sprint": "bat_sprint",
                # rank 33: the platoon-edge components (means of per-row
                # values computed below and attached to bdf)
                "lu_vsh_edge": "_vsh_tb_edge", "lu_vsh_kedge": "_vsh_k_edge",
                # v8 (2026-07-15): banded whiff + damage lineup views
                # (career damage reads — PD_SHRINK v8 note), mirroring the
                # vectorized lu aggregation exactly
                "lu_fblowh": "bd_fblowh_d", "lu_fbmidwh": "bd_fbmidwh_d",
                "lu_fb95wh": "bd_fb95wh_d",
                "lu_fbloxw": "bd_fbloxw_c", "lu_fbmidxw": "bd_fbmidxw_c",
                "lu_fb95xw": "bd_fb95xw_c",
                "lu_brkxw": "bd_brkxw_c", "lu_offxw": "bd_offxw_c",
                "lu_fbkxw": "bd_fbkxw_c"}

    def _starter_rows(self, spec, bdf=None, bmeta=None):
        """K-model rows. bdf/bmeta (the batter rows) supply the
        opposing-lineup aggregates; without them the lu_* features are NaN."""
        date = pd.Timestamp(spec["date"])
        season = date.year
        s = self.stores
        wx = self._weather(spec)
        park = s.park(spec["venue"], date)
        env = s.league_env(date)
        ump = s.ump_feats(spec.get("hp_ump_id"), date)
        rows, meta = [], []
        for side, pid, team, opp, home in [
                ("away", spec["away_starter"], spec["away_team"],
                 spec["home_team"], 0),
                ("home", spec["home_starter"], spec["home_team"],
                 spec["away_team"], 1)]:
            if not pid:  # starter not specified -> no K projection for side
                continue
            self._starter_sanity(pid, spec)
            f = s.starter_feats(pid, date, season, home=home)
            vs = s.team_offense(opp, season, date, prefix="vs")
            p_bio = s.bio(pid)
            row = {"Season": season, "month": date.month, "Home": home,
                   **f, **vs, **park, **wx, **env, **ump,
                   **s.pd_pitcher_feats(pid, date),
                   # his own TTO decay + the defense playing behind him
                   **s.tto(pid, date),
                   **s.lineup_oaa([p for p, _ in spec[f"{side}_lineup"]],
                                  season, prefix="own"),
                   # MiLB translated prior, allowed side (2026-07-13)
                   **s.milb_feats(pid, season, "pit"),
                   # 2026-07-14 finish batch: his contact quality allowed
                   # (the p_conv chain's damage half, #28), the manager's
                   # leash (#16), and his team's schedule context (#18/#24)
                   **s.bip_pitcher(pid, date),
                   "team_st_outs_pg": s.team_st_outs(team, season, date),
                   # audit wave: his OWN bullpen state (#6), own defense's
                   # unearned-run rate (#26), his SB-allowed permissiveness
                   # + league SB/27 centering (#24), his bio (#20), and his
                   # arsenal breadth/trajectory (#22)
                   **s.bullpen(team, season, date, prefix="own_pen"),
                   "own_pen_hl_era": s.bullpen_hl(
                       team, season, date, prefix="own_pen_hl")["own_pen_hl_era"],
                   "own_pen_np_l3": s.pen_fatigue(team, date),
                   "own_def_uer": s.team_defense(team, season, date),
                   **s.pitcher_prior(pid, season),
                   "lg_sb27_prior": s.lg_sb27_prior(season),
                   # battery + IL wave (2026-07-15): his own battery + his
                   # own IL return context
                   **s.catcher_feats(team, season, "own"),
                   **s.il_feats(pid, date, "p_"),
                   "pit_age": ((date - p_bio["dob"]).days / 365.25
                               if pd.notna(p_bio["dob"]) else np.nan),
                   "pit_height": p_bio["height"], "pit_weight": p_bio["weight"],
                   **s.arsenal_dynamics(pid, season),
                   **self._sched(spec, team)}
            F.add_pit_trends(row)
            # the starter's own arsenal, K-model view (same helper as training)
            pa = F.pitcher_arsenal_feats(
                pd.DataFrame({"PitcherId": [pid], "Season": [season]}),
                s.raw["ars_p"])
            for c in [*F.ARS_K_METRICS.values(), "pars_cov"]:
                row[c] = pa[c].iloc[0]
            # the actual opposing lineup he faces (mean of the batter rows)
            for dst, src in self._LU_COLS.items():
                if bdf is not None and bmeta is not None and len(bdf):
                    mask = (bmeta["Team"] == opp).to_numpy()
                    row[dst] = bdf.loc[mask, src].mean() if mask.any() else np.nan
                else:
                    row[dst] = np.nan
            rows.append(row)
            meta.append({"PlayerId": pid, "Team": team, "Opponent": opp,
                         "Name": self._name(pid, spec)})
        df = pd.DataFrame(rows)
        if len(df):
            # shared derived features (parity with build_starts_frame):
            # weather, TTO decay, K-BB, lineup-collision products
            df = F.add_starter_derived(df)
        return df, pd.DataFrame(meta)

    def _sim_game(self, spec):
        """Game-level PA-sim outputs for the SIM_BLEND layer, or None
        (missing artifacts, partial lineups, or a sim failure — every
        path degrades to the incumbent alone). Loads lazily, once."""
        if self._slate_sim is False:
            return None
        if self._slate_sim is None:
            try:
                from pa_serve import SlateSim
                self._slate_sim = SlateSim()
            except Exception as e:              # noqa: BLE001
                print(f"[predict] PA-sim unavailable ({e}); incumbent "
                      f"game heads alone")
                self._slate_sim = False
                return None
        try:
            return self._slate_sim.game(spec)
        except Exception as e:                  # noqa: BLE001
            print(f"[predict] PA-sim failed for this game ({e}); "
                  f"incumbent alone")
            return None

    def _starter_sanity(self, pid, spec):
        """Soft check: warn when a listed starter is a bullpen arm on the
        current roster (likely a mis-picked player in the GUI)."""
        try:
            pos = self.stores.rosters.loc[pid, "Position"]
        except KeyError:
            return
        if isinstance(pos, str) and pos.strip().lower() == "bullpen":
            print(f"note: listed starter {self._name(pid, spec)} is a "
                  f"bullpen arm on the current roster — check the pick")

    def _team_rows(self, spec, bdf=None, bmeta=None):
        """One row per team: own offense vs opposing pitching (order:
        away first, home second). bdf/bmeta (the batter rows) supply the
        posted-lineup aggregates (#19/#30/#32); without them those
        features are NaN."""
        date = pd.Timestamp(spec["date"])
        season = date.year
        s = self.stores
        wx = self._weather(spec)
        park = s.park(spec["venue"], date)
        env = s.league_env(date)
        ump = s.ump_feats(spec.get("hp_ump_id"), date)
        rows = []
        sides = [(spec["away_team"], spec["home_starter"],
                  spec["home_team"], 0, "away", "home"),
                 (spec["home_team"], spec["away_starter"],
                  spec["away_team"], 1, "home", "away")]
        for team, opp_starter, opp, home, own_side, opp_side in sides:
            opp_starter = opp_starter if opp_starter else -1
            own_lu = [p for p, _ in spec[f"{own_side}_lineup"]]
            opp_lu = [p for p, _ in spec[f"{opp_side}_lineup"]]
            toff = s.team_offense(team, season, date, prefix="off")
            toff_loc = s.team_offense_loc(team, season, home, date,
                                          prefix="off_loc")
            pen = s.bullpen(opp, season, date, prefix="opp_pen")
            pen_hl = s.bullpen_hl(opp, season, date, prefix="opp_pen_hl")
            stf = s.starter_feats(opp_starter, date, season)
            pb = s.bip_pitcher(opp_starter, date)
            lu = self._lineup_agg(bdf, bmeta, team)
            sc = self._sched(spec, team)
            rows.append({
                "Season": season, "month": date.month, "Home": home,
                "off_hr_pa": toff["off_hr_pa"], "off_r_pg": toff["off_r_pg"],
                "off_k_pct": toff["off_k_pct"],
                "off_loc_hr_pa": toff_loc["off_loc_hr_pa"],
                "off_loc_r_pg": toff_loc["off_loc_r_pg"],
                **s.team_bip_offense(team, season, date),
                "opp_pen_era": pen["opp_pen_era"],
                "opp_pen_hl_era": pen_hl["opp_pen_hl_era"],
                "opp_pen_np_l3": s.pen_fatigue(opp, date),
                "opp_def_uer": s.team_defense(opp, season, date),
                "opp_def_oaa": s.team_oaa(opp, season),
                "opp_pen_xwcon": s.team_bip_pen(opp, season,
                                                date)["pen_xwcon"],
                "opp_ps_era": stf["ps_era"], "opp_ps_k_bf": stf["ps_k_bf"],
                "opp_ps_hr_bf": stf["ps_hr_bf"], "opp_ps_h_bf": stf["ps_h_bf"],
                "opp_ps_xwcon": pb["pbip_xwoba"],
                "opp_ps_xwcon_d": pb["pbipd_xwoba"],
                "opp_ps_brl_d": pb["pbipd_brl"],
                "opp_ps_gb_d": pb["pbipd_gb"],
                "opp_pc_era": stf["pc_era"],
                "opp_pc_hr_bf": stf["pc_hr_bf"], **park, **wx, **env,
                # full-surface pass (2026-07-12): actual defense faced, own
                # lineup baserunning, opposing starter TTO sums (decay
                # derived below on the assembled frame, matching the
                # vectorized ps_tto_decay)
                **s.lineup_oaa(opp_lu, season, prefix="opp"),
                **s.lineup_brr(own_lu, season),
                **s.tto(opp_starter, date),
                # 2026-07-14 finish batch: B-lineup gap (#19), lineup air
                # profile + arsenal collision (#30/#32), BaseRuns luck
                # (#20), ump run environment (#21), schedule/travel + DH
                # flags (#18/#24)
                "off_lu_obp_gap": lu["lu_obp"] - toff["off_obp"],
                "off_lu_slg_gap": lu["lu_slg"] - toff["off_slg"],
                "off_lu_pullair": lu["lu_pullair"],
                "off_lu_arswh": lu["lu_arswh"],
                "off_bsr_luck": s.team_bsr(team, date),
                "ump_r_g": ump["ump_r_g"],
                "off_dan": sc["day_after_night"],
                "off_travel_km": sc["travel_km"],
                "off_tz_delta": sc["tz_delta"],
                "is_dh": sc["is_dh"], "dh_game2": sc["dh_game2"],
                # audit wave: walk channel (#8; patience_wild computed in
                # add_team_game_derived), starter length/leash + handoff gap
                # (#9; xpen_r_gap in the shared helper) + HL-outs share
                # rider, venue dispersion (#32), platoon edge (#33)
                "off_bb_pct": toff["off_bb_pct"],
                "opp_ps_bb_bf": stf["ps_bb_bf"],
                "lg_bb_pa": env["lg_bb_pa"],
                "opp_ps_len": stf["p_ip_per_start"],
                "opp_st_outs_pg": s.team_st_outs(opp, season, date),
                "opp_penhl_share": s.penhl_share(opp, season, date),
                "park_vmr": park["park_vmr"],
                "off_lu_vsh_edge": lu["lu_vsh_edge"],
                "off_lu_vsh_kedge": lu["lu_vsh_kedge"],
            })
        df = pd.DataFrame(rows)
        df["opp_ps_tto_decay"] = F.tto_decay_from_sums(df)
        # shared weather derivation (parity with build_game_frame ->
        # build_team_game_frame); the totals head reads hum_eff/air_dens
        # too, then the general carry wind (2026-07-14: was missing from
        # the serving team rows — served NaN; caught by the new team-row
        # selftest) and the shared team-game derived features (#30)
        F.add_wind_carry(df)
        return F.add_team_game_derived(F.add_weather_derived(df))

    @staticmethod
    def _lineup_agg(bdf, bmeta, team):
        """Posted-lineup aggregates from the served batter rows, mirroring
        features.lineup_aggregates (NaN-skipping means over the lineup)."""
        nan = {"lu_obp": np.nan, "lu_slg": np.nan, "lu_pullair": np.nan,
               "lu_arswh": np.nan, "lu_vsh_edge": np.nan,
               "lu_vsh_kedge": np.nan}
        if bdf is None or bmeta is None or not len(bdf):
            return nan
        mask = (bmeta["Team"] == team).to_numpy()
        if not mask.any():
            return nan
        return {"lu_obp": bdf.loc[mask, "_obpp"].mean(),
                "lu_slg": bdf.loc[mask, "_slgp"].mean(),
                "lu_pullair": bdf.loc[mask, "bip_pullair"].mean(),
                "lu_arswh": bdf.loc[mask, "arsenal_whiff"].mean(),
                # rank 33: the platoon edges (precomputed per batter row)
                "lu_vsh_edge": bdf.loc[mask, "_vsh_tb_edge"].mean(),
                "lu_vsh_kedge": bdf.loc[mask, "_vsh_k_edge"].mean()}

    # starter features the winner model consumes, per side
    _WIN_ST = ["ps_era", "ps_k_bf", "ps_hr_bf", "ps_bb_bf",
               "pc_era", "pc_hr_bf", "p_days_rest", "p5_k_bf"]

    def _win_row(self, spec, bdf=None, bmeta=None):
        """One row for the home-win classifier: both teams' records,
        offenses, bullpens, and starters, mirroring build_game_frame.
        bdf/bmeta supply the posted-lineup aggregates for the B-lineup
        gap diff (#19) and the arsenal-collision diff (#32)."""
        date = pd.Timestamp(spec["date"])
        season = date.year
        s = self.stores
        row = {"Season": season, "month": date.month,
               **s.park(spec["venue"], date), **self._weather(spec)}
        for side, team, starter in (("away", spec["away_team"],
                                     spec["away_starter"]),
                                    ("home", spec["home_team"],
                                     spec["home_starter"])):
            for k, v in s.team_record(team, season, date).items():
                row[f"{side}_{k}"] = v
            for k, v in s.team_form(team, season, date).items():
                row[f"{side}_{k}"] = v
            row[f"{side}_elo"] = s.team_elo(team, season, date)
            off = s.team_offense(team, season, date, prefix="off")
            row[f"{side}_r_pg"] = off["off_r_pg"]
            row[f"{side}_pen_era"] = s.bullpen(team, season, date)["pen_era"]
            stf = s.starter_feats(starter if starter else -1, date, season)
            for k in self._WIN_ST:
                row[f"{side}_{k}"] = stf[k]
            row[f"{side}_ps_xwcon_d"] = s.bip_pitcher(
                starter if starter else -1, date)["pbipd_xwoba"]
            # 2026-07-14 finish batch: per-side BaseRuns luck, B-lineup
            # gap, arsenal collision (diffs below feed the winner)
            lu = self._lineup_agg(bdf, bmeta, team)
            row[f"{side}_bsr_luck"] = s.team_bsr(team, date)
            row[f"{side}_lu_obp_gap"] = lu["lu_obp"] - off["off_obp"]
            row[f"{side}_lu_arswh"] = lu["lu_arswh"]
        for f in ["win_pct", "rd_pg", "pyth", "ra_pg", "r_pg", "w20", "rd20",
                  "ps_era", "pc_era", "pen_era", "ps_k_bf", "ps_xwcon_d",
                  "bsr_luck", "lu_obp_gap", "lu_arswh"]:
            row[f"d_{f}"] = row[f"home_{f}"] - row[f"away_{f}"]
        row["d_rest"] = row["home_p_days_rest"] - row["away_p_days_rest"]
        row["d_ps_kbb"] = ((row["home_ps_k_bf"] - row["home_ps_bb_bf"])
                           - (row["away_ps_k_bf"] - row["away_ps_bb_bf"]))
        row["d_elo"] = row["home_elo"] - row["away_elo"]
        row["elo_prob_home"] = F.elo_expected(row["home_elo"], row["away_elo"])
        return row

    # ------------------------------------------------------- predict

    def _prep(self, df, cols):
        # audit #12: a serving column the row-builders failed to produce
        # used to become silent NaN ("imputed missing"). Still filled — a
        # missing optional source must not kill a slate — but now WARNED,
        # once per column per session, so train/serve drift is visible.
        missing = [c for c in cols if c not in df.columns]
        new = [c for c in missing
               if c not in getattr(self, "_warned_missing", set())]
        if new:
            self._warned_missing = getattr(self, "_warned_missing",
                                           set()) | set(new)
            print(f"WARNING: serving produced no value for {len(new)} model "
                  f"column(s) — filled NaN (train/serve drift?): "
                  f"{new[:8]}{'...' if len(new) > 8 else ''}")
        for c in missing:
            df[c] = np.nan
        df = df[cols].copy()
        for c, levels in self.art["cat_levels"].items():
            if c in df.columns:
                df[c] = pd.Categorical(df[c].astype("object"), categories=levels)
        for c in df.columns:
            if c not in self.art["cat_levels"]:
                df[c] = pd.to_numeric(df[c], errors="coerce")
        return df

    def predict_game(self, spec):
        a = self.art
        if not spec.get("away_lineup") and not spec.get("home_lineup"):
            raise ValueError("at least one lineup player is required")
        bdf, bmeta = self._batter_rows(spec)
        X = self._prep(bdf, a["bat_cols"])

        batters = bmeta[["Team", "slot", "PlayerId", "Name"]].copy()
        # G# (audit #10): game number within the day (2 = doubleheader game
        # 2), so the grader scores each row against ITS game's box line
        # instead of the day sum
        batters.insert(1, "G#", 2 if spec.get("dh_game2") else 1)
        # career-games flag: picks on players with under ~50 MLB games were
        # the weakest segment in the holdout eval (AUC 0.58 vs 0.63 overall)
        batters["CareerG"] = pd.to_numeric(
            bdf["g_career"], errors="coerce").fillna(0).astype(int).to_numpy()
        offs = self.offsets if self.recal else {}
        # two passes: raw probabilities first, so stacked props (see
        # apply_stack) can see their donors' scores
        raw_p = {prop: predict_prop(a["props"][prop], X)
                 for prop in PROP_COLS if prop in a["props"]}
        fin_p = {}
        for prop in raw_p:
            p = apply_stack(a["props"][prop], raw_p[prop], raw_p)
            if offs.get(prop):
                p = R.apply_offset(p, offs[prop])
            fin_p[prop] = p
        # threshold-ladder coherence (2026-07-15): P(3+ TB) may never exceed
        # P(2+ TB) etc. — same projection evaluate_deep applies, so the
        # served prices are the ones the paired read verdicts
        fin_p = F.enforce_ladders(fin_p)
        for prop, col in PROP_COLS.items():
            if prop not in fin_p:  # artifact predates this prop
                continue
            batters[col] = np.round(fin_p[prop], 4)
        batters["xHR"] = np.round(batters["P_HR"] * a["multi_hr"], 4)
        batters["HR_fair_odds"] = [american_odds(p) for p in batters["P_HR"]]
        # count-head means (xSO = batter Ks, xHRR = hits+runs+RBIs,
        # xTB = total bases)
        for cname, col in BAT_COUNT_COLS.items():
            head = a.get("count_models", {}).get(cname)
            if head is not None:
                batters[col] = np.round(
                    head["model"].predict(self._prep(bdf, head["cols"])), 2)
        batters = batters.sort_values("P_HR", ascending=False).reset_index(drop=True)

        sdf, smeta = self._starter_rows(spec, bdf, bmeta)
        if len(sdf):
            Xs = self._prep(sdf, a["st_cols"])
            k_pred = a["k_model"].predict(Xs)
        else:
            k_pred = np.array([])
        starters = smeta if len(smeta) else pd.DataFrame(
            columns=["PlayerId", "Team", "Opponent", "Name"])
        starters = starters.copy()
        starters["G#"] = 2 if spec.get("dh_game2") else 1   # audit #10
        starters["xK"] = np.round(k_pred, 2)
        # K P(over): cal-year per-line logistic (k_over), NB/Poisson fallback
        for line in K_LINES:
            starters[f"P_over_{line}"] = np.round(
                k_over(a, k_pred, line), 3).tolist()
        # starter count props: outs recorded, walks/hits allowed (mean +
        # NB P(over) with each head's calibration-year dispersion)
        if len(sdf):
            for cname, (mcol, pre) in ST_COUNT_COLS.items():
                head = a.get("count_models", {}).get(cname)
                if head is None:
                    continue
                mu = head["model"].predict(self._prep(sdf, head["cols"]))
                starters[mcol] = np.round(mu, 2)
                for line in head["lines"]:
                    starters[f"{pre}_{line}"] = np.round(
                        count_over(head, mu, line), 3)
        gdf = self._team_rows(spec, bdf, bmeta)
        Xg = self._prep(gdf, a["tg_cols"])
        mu_away, mu_home = a["team_runs_model"].predict(Xg)
        total_runs = float(mu_away + mu_home)
        # winner: the dedicated home-win model (team strength/Elo/form +
        # both starters, blended with the Poisson win prob); the bare
        # Poisson comparison is only the fallback for old artifacts
        if "win_model" in a:
            Xw = self._prep(pd.DataFrame([self._win_row(spec, bdf, bmeta)]),
                            a["win_model"]["cols"])
            p_home = float(predict_win(a["win_model"], Xw,
                                       mu_home, mu_away)[0])
        else:
            p_home = poisson_win(mu_home, mu_away)
        # PA-sim game-level blend (Phase 3, 2026-07-13): one sim of this
        # game; score/total means blend linearly, the winner on logits, at
        # the fixed SIM_BLEND weights. None -> incumbent alone.
        sim = self._sim_game(spec) if SIM_BLEND else None
        if sim is not None:
            w = SIM_BLEND.get("score", 0.0)
            mu_away = (1 - w) * float(mu_away) + w * sim["x_away"]
            mu_home = (1 - w) * float(mu_home) + w * sim["x_home"]
            w = SIM_BLEND.get("total", 0.0)
            total_runs = (1 - w) * total_runs + w * sim["x_total"]
            w = SIM_BLEND.get("winner", 0.0)
            if w:
                z = (w * _sim_logit(sim["p_home_win"])
                     + (1 - w) * _sim_logit(p_home))
                p_home = float(1 / (1 + np.exp(-z)))
        # Home-win PROBABILITY, not a pick: on the holdout the winner model
        # has no statistically significant edge over always taking the home
        # team, so it's presented as a calibrated number, never a bet.
        totals = {
            "exp_lineup_HR": round(float(batters["xHR"].sum()), 2),
            "exp_away_runs": round(float(mu_away), 2),
            "exp_home_runs": round(float(mu_home), 2),
            "exp_total_runs": round(total_runs, 2),
            "home_team": spec["home_team"], "away_team": spec["away_team"],
            "home_win_prob": round(float(p_home), 3),
            "P_over_runs": {str(l): round(float(total_over(
                a, total_runs, l,
                park_vmr=self.stores.park(
                    spec["venue"], pd.Timestamp(spec["date"]))["park_vmr"])[0]),
                3)
                            for l in TOTAL_LINES},
            # H5 team_total (2026-07-14): per-team lines off the SIM-blended
            # per-team means — coherent with the displayed expected scores
            "P_over_away_runs": {str(l): round(float(
                team_over(a, mu_away, l)[0]), 3) for l in TEAM_TOTAL_LINES},
            "P_over_home_runs": {str(l): round(float(
                team_over(a, mu_home, l)[0]), 3) for l in TEAM_TOTAL_LINES},
        }
        return {"batters": batters, "starters": starters, "totals": totals}

    def predict_slate(self, specs):
        """Predict several games; batter/starter boards are combined and the
        HR board is ranked across the whole slate."""
        all_b, all_s, games = [], [], []
        for spec in specs:
            out = self.predict_game(spec)
            tag = f'{spec["away_team"]}@{spec["home_team"]}'
            b = out["batters"].copy()
            b.insert(0, "Game", tag)
            s = out["starters"].copy()
            s.insert(0, "Game", tag)
            t = out["totals"]
            p_home = t["home_win_prob"]
            fav_home = p_home >= 0.5
            games.append({"Game": tag, "Date": spec["date"],
                          "Venue": spec["venue"],
                          "Winner": t["home_team"] if fav_home else t["away_team"],
                          "WinProb": p_home if fav_home else round(1 - p_home, 3),
                          "HomeWinProb": t["home_win_prob"],
                          "exp_away_runs": t["exp_away_runs"],
                          "exp_home_runs": t["exp_home_runs"],
                          "exp_lineup_HR": t["exp_lineup_HR"],
                          "exp_total_runs": t["exp_total_runs"],
                          **{f"P_runs_over_{k}": v
                             for k, v in t["P_over_runs"].items()},
                          **{f"P_away_runs_over_{k}": v
                             for k, v in t["P_over_away_runs"].items()},
                          **{f"P_home_runs_over_{k}": v
                             for k, v in t["P_over_home_runs"].items()}})
            all_b.append(b)
            all_s.append(s)
        return {
            "batters": pd.concat(all_b).sort_values(
                "P_HR", ascending=False).reset_index(drop=True),
            "starters": pd.concat(all_s).reset_index(drop=True),
            "games": pd.DataFrame(games),
        }


# -------------------------------------------------------------- reporting

PRED_DIR = Path(__file__).resolve().parents[1] / "Predictions"

# Excel-facing headers. Internal frames keep the raw P_*/x* names (the GUI
# and summary_frame read those); the workbook gets short display names.
BAT_HEADERS = {"slot": "Slot", "PlayerId": "ID", "CareerG": "Career G",
               "P_HR": "HR", "P_Hit": "Hit", "P_2Hits": "2+ Hits",
               "P_1B": "Single", "P_2B": "Double", "P_3B": "Triple",
               "P_TB2": "2+ TB", "P_TB3": "3+ TB", "P_TB4": "4+ TB",
               "P_Run": "Run", "P_Run2": "2+ Runs",
               "P_RBI": "RBI", "P_RBI2": "2+ RBI",
               "P_HRR2": "H+R+RBI 2+", "P_HRR3": "H+R+RBI 3+",
               "P_HRR4": "H+R+RBI 4+",
               "P_BB": "BB", "P_SB": "SB",
               "P_K": "K", "P_2K": "2+ K", "P_3K": "3+ K"}
ST_HEADERS = {"PlayerId": "ID"}
GAME_HEADERS = {"WinProb": "Win Prob",
                "exp_away_runs": "Away Score", "exp_home_runs": "Home Score",
                "exp_total_runs": "Total Runs", "exp_lineup_HR": "Lineup HRs"}

# over-probability columns are renamed by pattern (any line value works)
_OVER_PATTERNS = [(re.compile(r"^P_over_(.+)$"), "K > {}"),
                  (re.compile(r"^P_outs_over_(.+)$"), "Outs > {}"),
                  (re.compile(r"^P_bb_over_(.+)$"), "BB > {}"),
                  (re.compile(r"^P_hits_over_(.+)$"), "Hits > {}"),
                  (re.compile(r"^P_er_over_(.+)$"), "ER > {}"),
                  (re.compile(r"^P_away_runs_over_(.+)$"), "Away Runs > {}"),
                  (re.compile(r"^P_home_runs_over_(.+)$"), "Home Runs > {}"),
                  (re.compile(r"^P_runs_over_(.+)$"), "Runs > {}")]

# reading order per sheet; an entry ending in "> " pulls that whole family
# of over columns, sorted by line
BAT_ORDER = ["Game", "Team", "G#", "Slot", "Name", "ID", "Career G",
             "HR",
             "xTB", "2+ TB", "3+ TB", "4+ TB",
             "xH", "Hit", "2+ Hits",
             "xRBI", "RBI", "2+ RBI",
             "xR", "Run", "2+ Runs",
             "xHRR", "H+R+RBI 2+", "H+R+RBI 3+", "H+R+RBI 4+",
             "SB",
             "xSO", "K", "2+ K", "3+ K",
             "Single",
             "Double",
             "Triple",
             "xBB", "BB"]
ST_ORDER = ["Game", "Team", "Opponent", "G#", "Name", "ID",
            "xK", "K > ",
            "xER", "ER > ",
            "xOuts", "Outs > ",
            "xHits", "Hits > ",
            "xBB", "BB > "]
GAME_ORDER = ["Game", "Date", "Venue", "Winner", "Win Prob",
              "Away Score", "Away Runs > ",
              "Home Score", "Home Runs > ",
              "Total Runs", "Runs > ",
              "Lineup HRs"]


def _display(df, headers, order, drop=()):
    """Workbook view of a board: drop internal-only columns, give every
    header a short display name, and arrange columns in reading order.
    Columns the order list doesn't know about keep their place at the end."""
    df = df.drop(columns=[c for c in drop if c in df.columns])
    ren = {}
    for c in df.columns:
        if c in headers:
            ren[c] = headers[c]
        else:
            for pat, fmt in _OVER_PATTERNS:
                m = pat.match(c)
                if m:
                    ren[c] = fmt.format(m.group(1))
                    break
    df = df.rename(columns=ren)
    cols, out = list(df.columns), []
    for item in order:
        if item.endswith("> "):
            out += sorted((c for c in cols if c.startswith(item)),
                          key=lambda c: float(c.rsplit(" ", 1)[1]))
        elif item in cols:
            out.append(item)
    out += [c for c in cols if c not in out]
    return df[out]

GLOSSARY = [
    ("How to read this workbook",
     "Every number is the model's estimated chance that something happens "
     "in this game. 25% means: in 100 similar situations, expect it about "
     "25 times. Nothing is ever certain."),
    ("Highlighted columns", "The red-tinted columns are the headline "
     "numbers: the model's expected counts (xK, xOuts, xHits, xBB, xER, "
     "xSO, xHRR, xTB), and the predicted winner, win probability, total runs "
     "and "
     "lineup HRs on the Games sheet. The percentage columns next to the "
     "counts price the over/under lines."),
    ("HR", "Chance the batter hits a home run in this game."),
    ("Hit", "Chance of at least one hit."),
    ("2+ Hits", "Chance of two or more hits."),
    ("Single", "Chance of at least one single."),
    ("Double", "Chance of at least one double."),
    ("Triple", "Chance of at least one triple — the rarest hit on the "
     "board (about 1 batter-game in 80), so even the best numbers here "
     "are small."),
    ("xTB", "Expected total bases (1B + 2x2B + 3x3B + 4xHR) — the headline "
     "count behind the total-bases market."),
    ("2+ TB / 3+ TB / 4+ TB", "Chance of two / three / four or more total "
     "bases (the total-bases prop at the 1.5 / 2.5 / 3.5 line)."),
    ("xH", "Expected hits."),
    ("Run / 2+ Runs", "Chance the batter scores at least one / two runs."),
    ("xR", "Expected runs scored."),
    ("RBI / 2+ RBI", "Chance the batter drives in at least one / two runs."),
    ("xRBI", "Expected runs batted in."),
    ("xHRR", "Expected combined hits + runs + RBIs."),
    ("H+R+RBI 2+ / 3+ / 4+", "Chance of 2+ / 3+ / 4+ combined hits + runs "
     "+ RBIs (the H+R+RBI prop at the 1.5 / 2.5 / 3.5 line)."),
    ("BB", "Chance of at least one walk."),
    ("xBB (Batter sheet)", "Expected walks drawn by the batter."),
    ("SB", "Chance of at least one stolen base."),
    ("xSO", "Expected strikeouts by the batter."),
    ("K / 2+ K / 3+ K", "Chance the batter strikes out at least once / "
     "twice / three times."),
    ("Career G", "The batter's career MLB games before today. Predictions "
     "for players under ~50 games are the least reliable (the model has "
     "little history to work from) - those show in red; treat their picks "
     "with extra caution."),
    ("ID", "The player's MLB id, for cross-referencing stats sites; safe "
     "to ignore."),
    ("xK", "Projected strikeouts for the starting pitcher."),
    ("K > X", "Chance the starter records more than X strikeouts."),
    ("xOuts", "Projected outs recorded by the starter (18 = six innings)."),
    ("Outs > X", "Chance the starter records more than X outs."),
    ("xBB / BB > X", "Projected walks allowed by the starter, and the "
     "chance of more than X."),
    ("xHits / Hits > X", "Projected hits allowed by the starter, and "
     "the chance of more than X."),
    ("xER / ER > X", "Projected earned runs allowed by the starter, and "
     "the chance of more than X."),
    ("Expected counts vs. their over/under", "The x-counts (xK, xOuts, "
     "xHits, xBB, xER) are AVERAGES. Because a few blow-up innings pull the "
     "average up while most starts come in lower, the 50/50 point of the "
     "over/under sits a little BELOW the average - so e.g. an xHits of 4.6 "
     "can still be under 50% to go over 4.5. The two columns are the same "
     "prediction shown two ways, not a contradiction."),
    ("Fair odds (Summary sheet)", "The break-even sportsbook price for the "
     "HR chance. If a book offers longer odds (bigger + number), the bet "
     "pays more than the risk suggests; shorter odds pay less."),
    ("Lineup HRs", "Expected total home runs by the players entered."),
    ("Total Runs / Runs > X", "Expected combined runs scored by both "
     "teams, and the chance of more than X."),
    ("Away/Home Runs > X", "Chance ONE team scores more than X runs (the "
     "team-total market), priced off the same expected scores shown in "
     "Away Score / Home Score."),
    ("Winner / Win Prob", "The team the model favors and its win "
     "probability (always the bigger side of the home/away split). Treat it "
     "as a probability, not a pick: on a half-season holdout the winner "
     "model shows no statistically significant edge over always taking the "
     "home team."),
    ("Slot", "Batting-order position (1 = leadoff)."),
    ("G#", "Game number within the day: 1 normally, 2 for the second game "
     "of a doubleheader. The grader uses it to score each row against its "
     "own game's box line instead of the day's combined stats."),
    ("Bets sheet", "The model's betting signals for this slate: every side "
     "where the model's probability beats the best posted sportsbook price by "
     "an expected 5%+ per $1 staked, after removing the book's margin (vig). "
     "Sorted by EV — the top row is the strongest edge. If it's empty or shows "
     "a note, nothing cleared the bar (or no odds were captured — run "
     "Tools/2_scrape_odds.py near game time)."),
    ("Bets columns", "Model % is the model's chance for the side shown; Mkt % "
     "is the de-vigged market chance; Edge is their difference; Best Odds / "
     "Book are the most generous posted American price and the book offering "
     "it; EV% is the expected profit per $1 at that price. Note flags a rookie "
     "(<50 career games) or a thin one-book price."),
    ("Green cells", "A green cell on the Batter/Pitching/Games sheets is a "
     "flagged OVER bet — the model prices the 'over' side (the stat clears "
     "the line; for a game it is more runs, or the home team on the Win Prob "
     "cell) at >=5% EV against the best posted price, shown where its number "
     "lives so you can see it in context. UNDER bets are just as real but are "
     "NOT painted here — the displayed cell is the over probability, so a "
     "green under would point at the wrong number; every under (with its Side) "
     "is on the Bets sheet, which carries both directions."),
    ("Blue cells", "A light-blue cell is a rank-quality pick: on the model's "
     "own held-out diagnostics (the PROP_RANKINGS playbook, computed fresh "
     "from the current model — never the workbook file), that column has "
     "PROVEN selection power and this row is one of today's best picks for "
     "it. Everything is measured with a confidence interval, not a lucky "
     "point estimate: the column earns depth only if the LOWER BOUND of its "
     "odds-ratio top-pick lift clears the bar (base-rate-fair, so a "
     "high-base column like Hit competes on equal footing) and its "
     "calibration slope is sane; the depth is then capped by the tier of its "
     "Score_lo — the day-block-bootstrap LOWER bound of the rankings Score "
     "(for O/U columns, a 50/50 blend of the market row and that line's own "
     "row, so both PROP_RANKINGS tables vote): up to 10 cells for STRONG or "
     "better, 7 for SOLID, 5 for DECENT, 4 for LOW CEILING, 2 below that, "
     "across the whole slate. So a THIN market (a rare prop, a deep line) "
     "carries a wide "
     "interval, a lower Score_lo, and shallower blue — it can no longer buy "
     "depth on one good season. Marks are OVER-side only, ranked by "
     "probability, and each must clear an INFORMEDNESS FLOOR over that "
     "column's own base rate — a likely line is never marked just for being "
     "likely (K > 3.5, which hits ~67% of the time, needs 77.5%+ to be "
     "painted; K > 8.5 needs 15.4%+). When sportsbook lines exist for the "
     "exact market the sharp consensus acts as a VETO only (a pick the sharp "
     "books price at under half the model's number is dropped) — odds never "
     "SELECT a blue mark; value selection stays with the green cells. A pick "
     "that is BOTH a rank-quality pick and a +EV bet shows light purple — "
     "the strongest signal on the sheet."),
    ("Graded colors", "After the games, Tools/4_grade_results.py re-colors "
     "the workbook from the actual box scores (our own scraped data, "
     "matched by player ID). The grammar: DARK fill + white text = the "
     "pick HIT (dark blue / dark green / dark purple); GRAYED-OUT = the "
     "pick missed; YELLOW = the stat occurred where nothing was picked; "
     "white = nothing picked, nothing happened. Every graded cell answers "
     "the literal question: did the stat occur — O/U cells light only if "
     "the OVER hit, the Winner cell only if the named team won. "
     "Re-running the script repairs and re-grades safely."),
    ("Bold probabilities", "Any probability above 50% is bold; at or "
     "below 50% is regular weight — no exceptions, including inside "
     "pick fills and graded cells (a HIT below 50% shows white regular "
     "text on its dark fill). Text on a fill (a picked Winner) stays "
     "bold for contrast. Misses are gray italic, bold only above 50%."),
    ("EV is not a guarantee", "The edge is computed from the model's OWN "
     "probability. Beating a naive base rate is not the same as beating a "
     "sharp book, and until many days of closing lines accrue these edges are "
     "unproven — treat small ones as noise and size same-game picks as one "
     "bet. Bet responsibly."),
    ("A caution", "These are model estimates from historical data. They do "
     "not account for injuries, breaking news, or the sportsbook's own "
     "information. Bet responsibly."),
]


def _top_str(df, col, n=3, extra=None):
    rows = df.nlargest(n, col)
    parts = []
    for _, r in rows.iterrows():
        s = f'{r["Name"]} ({r[col]:.0%}'
        if extra:
            s += f', fair odds {r[extra]}'
        parts.append(s + ")")
    return ";  ".join(parts)


def summary_frame(specs, out):
    """Plain-English 'what the model thinks' sheet, one block per game."""
    lines = []
    b_all, s_all = out["batters"], out["starters"]
    for i, spec in enumerate(specs):
        tag = f'{spec["away_team"]}@{spec["home_team"]}'
        b = b_all[b_all["Game"] == tag] if "Game" in b_all.columns else b_all
        s = s_all[s_all["Game"] == tag] if "Game" in s_all.columns else s_all
        t = _slate_totals(out, i) if "games" in out else out["totals"]
        lines += [
            (f'{spec["away_team"]} @ {spec["home_team"]}',
             f'{spec["date"]} — {spec["venue"] or "stadium TBD"}'),
            ("Most likely to homer", _top_str(b, "P_HR", 3, "HR_fair_odds")),
            ("Most likely to get a hit", _top_str(b, "P_Hit")),
            ("Most likely to score a run", _top_str(b, "P_Run")),
            ("Most likely to drive in a run", _top_str(b, "P_RBI")),
            ("Most likely to walk", _top_str(b, "P_BB")),
            ("Stolen base watch", _top_str(b, "P_SB")),
        ]
        for _, r in s.iterrows():
            lines.append((f'{r["Name"]} strikeouts',
                          f'projected {r["xK"]:.1f} — over 4.5: '
                          f'{r["P_over_4.5"]:.0%}, over 5.5: '
                          f'{r["P_over_5.5"]:.0%}, over 6.5: '
                          f'{r["P_over_6.5"]:.0%}'))
        hwp = t.get("home_win_prob", 0)
        lines.append(("Home win probability",
                      f'{spec["home_team"]} {hwp:.0%} vs '
                      f'{spec["away_team"]} {1 - hwp:.0%} — a calibrated '
                      f'estimate, not a pick (no proven edge over always '
                      f'taking the home team). Expected score '
                      f'{spec["away_team"]} {t.get("exp_away_runs", "?")}, '
                      f'{spec["home_team"]} {t.get("exp_home_runs", "?")}'))
        lines.append(("Game total",
                      f'expect about {t["exp_total_runs"]} combined runs '
                      f'(over 8.5 runs: {t["P_over_runs"]["8.5"]:.0%}) and '
                      f'{t["exp_lineup_HR"]} home runs from these lineups'))
        lines.append(("", ""))
    if "Game" in b_all.columns and len(specs) > 1:
        top = b_all.nlargest(min(15, len(b_all)), "P_HR")
        counts = top["Game"].value_counts()
        big_game, big_n = counts.index[0], int(counts.iloc[0])
        # effective independent picks if same-game picks are treated as one
        eff = int(round(counts.shape[0]))
        lines.append(("Slate exposure (HR board)",
                      f"the top {len(top)} HR picks span {counts.shape[0]} of "
                      f"{len(specs)} games; {big_game} supplies {big_n} of them. "
                      f"Same-game picks share park, weather and pitcher — they "
                      f"win and lose together, so size a game's picks as roughly "
                      f"ONE bet (~{eff} independent edges here, not {len(top)}) "
                      f"and cap combined stake per game."))
    return pd.DataFrame(lines, columns=["What", "The model says"])


# display headers holding a probability (shown as a percent); over columns
# ("K > 4.5" etc.) are recognized by the " > " in their name
PCT_COLS = {"HR", "Hit", "2+ Hits", "Single", "Double", "Triple",
            "2+ TB", "3+ TB", "4+ TB", "Run", "2+ Runs",
            "RBI", "2+ RBI", "H+R+RBI 2+", "H+R+RBI 3+", "H+R+RBI 4+",
            "BB", "SB", "K", "2+ K", "3+ K",
            "Win Prob"}
# the headline columns get the red highlight: the expected counts, and the
# Games sheet's predicted winner, win probability, total runs and lineup HRs


# ------------------------------------------------------- betting signals
#
# The prediction Excel now compares the model's live probabilities to the real
# sportsbook lines in the odds store (Data/mlb_odds.csv) and flags the +EV
# picks. This reuses the EXACT de-vig/EV math that evaluate_deep Section 9
# grades the model with (Model/odds.py) — the only difference is it runs on
# today's UNSETTLED slate instead of finished games. Nothing here feeds the
# model; odds are consulted at output time only.

# A side is flagged as a "bet" (highlighted green) when the model prices its
# expected profit per 1u staked at >= this, graded at the best posted book
# price. 5% keeps the README's vig/noise caveat from turning razor-thin,
# unproven edges into "bets".
BET_EV_THRESHOLD = 0.05

# model P(over line) column prefix for each pitcher count market key
STARTER_PREFIX = {"pk": "P_over_", "pouts": "P_outs_over_",
                  "phits": "P_hits_over_", "pbb": "P_bb_over_",
                  "per": "P_er_over_"}

BET_HEADERS = ["Game", "Player", "Team", "Prop", "Side", "Line",
               "Model %", "Mkt %", "Edge", "Best Odds", "Book", "EV%",
               "Books", "Note"]
BET_PCT_COLS = {"Model %", "Mkt %", "Edge", "EV%"}      # shown as percents
BET_TEXT_COLS = {"Player", "Prop", "Side", "Book", "Note", "Bets"}


def _as_int(x):
    try:
        return int(x)
    except (TypeError, ValueError):
        return None


def _over_display(col):
    """Display header for a raw P(over) column, matching _display's renames
    ('P_over_6.5' -> 'K > 6.5', 'P_er_over_2.5' -> 'ER > 2.5'); None if the
    column is not an over column."""
    for pat, fmt in _OVER_PATTERNS:
        m = pat.match(col)
        if m:
            return fmt.format(m.group(1))
    return None


def _consensus(store, api, line, key):
    """Per join `key` ('PlayerId' for player props, 'Team' = home club for game
    markets): de-vig every book's two-sided price for one market/line to a fair
    P(over), then keep the reference fair prob (Pinnacle's where it posts the
    line, else the median — odds.sharp_fair), the median hold, and the best
    (most generous) over/under price with the book offering it. Mirrors
    evaluate_deep Section 9's _market_consensus/_game_consensus so live signals
    use the same math the backtest grades. Returns {key_value: {...}}; empty if
    nothing prices."""
    m = store[store["Market"] == api]
    if line is not None:
        m = m[(m["Line"] - line).abs() < 1e-6]
    recs = []
    for _, r in m.iterrows():
        fo, hold = O.devig_two_way(r["OverPrice"], r["UnderPrice"])
        if np.isnan(fo):
            continue
        recs.append((r[key], fo, hold, r["OverPrice"], r["UnderPrice"],
                     r["Book"]))
    if not recs:
        return {}
    md = pd.DataFrame(recs, columns=[key, "fair", "hold", "over", "under",
                                     "book"])
    out = {}
    for kval, g in md.groupby(key, dropna=False):
        if pd.isna(kval):
            continue
        over, under = g["over"].dropna(), g["under"].dropna()
        out[_as_int(kval) if key == "PlayerId" else kval] = {
            "fair": O.sharp_fair(g, book_col="book"),
            "hold": (float(g["hold"].median()) if g["hold"].notna().any()
                     else float("nan")),
            "best_over": float(over.max()) if len(over) else float("nan"),
            "best_under": float(under.max()) if len(under) else float("nan"),
            "over_book": str(g.loc[over.idxmax(), "book"]) if len(over) else "",
            "under_book": (str(g.loc[under.idxmax(), "book"]) if len(under)
                           else ""),
            "n_books": int(g["book"].nunique()),
        }
    return out


def _bet_rec(game, player, team, prop, side, line, p, c, ev, note,
             side_label=None):
    """One Bets-sheet row. `side` is 'over'/'under' as O.pick_side returns it;
    `side_label` overrides the displayed side (e.g. a team for the moneyline).
    p is the model's P(over line); c is the _consensus record."""
    p_side = p if side == "over" else 1.0 - p
    mkt_side = c["fair"] if side == "over" else 1.0 - c["fair"]
    price = c["best_over"] if side == "over" else c["best_under"]
    book = c["over_book"] if side == "over" else c["under_book"]
    return {
        "Game": game, "Player": player, "Team": team or "", "Prop": prop,
        "Side": side_label or ("Over" if side == "over" else "Under"),
        "Line": "" if line is None else line,
        "Model %": round(p_side, 4), "Mkt %": round(mkt_side, 4),
        "Edge": round(p_side - mkt_side, 4),
        "Best Odds": int(round(price)) if pd.notna(price) else "",
        "Book": book, "EV%": round(ev, 4),
        "Books": c["n_books"], "Note": note,
    }


def _game_records(out, specs):
    """Normalize the game-level predictions (totals + moneyline) to a common
    shape for both the slate ('games' frame) and single-game ('totals' dict)
    paths. Home/away come from the specs so it works even when 'totals' omits
    the team names."""
    if "games" in out:
        g = out["games"]
        tcols = [c for c in g.columns if c.startswith("P_runs_over_")]
        recs = []
        for _, r in g.iterrows():
            away, home = str(r["Game"]).split("@")
            recs.append({"tag": r["Game"], "home": home, "away": away,
                         "home_win_prob": float(r["HomeWinProb"]),
                         "totals": {c[len("P_runs_over_"):]: r[c]
                                    for c in tcols}})
        return recs
    s, t = specs[0], out["totals"]
    home, away = s["home_team"], s["away_team"]
    return [{"tag": f"{away}@{home}", "home": home, "away": away,
             "home_win_prob": float(t.get("home_win_prob", 0.5)),
             "totals": dict(t.get("P_over_runs", {}))}]


def compute_bets(out, specs, store_path=None, ev_threshold=BET_EV_THRESHOLD):
    """Compare the model's live probabilities to the odds store and return the
    +EV betting board plus the cells to highlight green in the prop sheets.

    For every model probability (batter props, pitcher count-prop lines, game
    totals, moneyline) we look up the market via Model/odds.py's PROP_MARKET /
    STARTER_MARKET, de-vig the posted prices to a fair number, and let
    O.pick_side pick the +EV side at the best book price. A row is a "bet" when
    its expected profit per 1u >= ev_threshold.

    Returns (bets_df, highlights, note):
      bets_df    - one row per flagged bet, sorted by EV (BET_HEADERS columns),
                   BOTH sides — overs and unders
      highlights - {sheet_title: [(row_match, display_column), ...]} for _polish,
                   OVER-side bets only (the displayed cell describes the over,
                   so under bets are not painted on the grid — they still ship
                   on the Bets sheet with their Side)
      note       - a human message when there is nothing to show, else None
    """
    store = O.load_odds(store_path or O.DEFAULT_STORE)
    dates = {str(s["date"]) for s in specs}
    if not store.empty:
        store = store[store["Date"].astype(str).isin(dates)]
    if store.empty:
        return (pd.DataFrame(columns=BET_HEADERS), {},
                f"No sportsbook odds in the store for "
                f"{', '.join(sorted(dates))}. Run  python "
                f"Tools/2_scrape_odds.py  near game time to capture lines, "
                f"then re-run this prediction.")

    cache = {}

    def cons(api, line, key):
        ck = (api, line, key)
        if ck not in cache:
            cache[ck] = _consensus(store, api, line, key)
        return cache[ck]

    bets = []
    highlights = defaultdict(list)

    def add(rec, sheet, match, disp_col, side):
        bets.append(rec)
        # Prop-grid cells are painted green for the OVER only — the standard
        # "this happens" pick that the displayed number describes. Under bets
        # are real +EV plays and still appear on the Bets sheet (with their
        # Side), but painting the over-probability cell green for an under is
        # misleading (a green 11% HR cell that actually means "no HR"). Over
        # semantics per market: batter/pitcher props = the stat clears the
        # line; game total = more runs; moneyline = the home team (the side
        # the displayed Win Prob is for).
        if disp_col and side == "over":
            highlights[sheet].append((match, disp_col))

    single_tag = _game_records(out, specs)[0]["tag"] if specs else ""

    # ---- batter binary props (join on PlayerId) ----
    batters = out["batters"]
    b_game = "Game" in batters.columns
    for prop, pcol in PROP_COLS.items():
        meta = O.PROP_MARKET.get(prop)
        if meta is None or pcol not in batters.columns:
            continue
        c_by_pid = cons(meta["api"], meta["line"], "PlayerId")
        if not c_by_pid:
            continue
        disp = BAT_HEADERS.get(pcol, pcol)
        for _, r in batters.iterrows():
            pid = _as_int(r.get("PlayerId"))
            c = c_by_pid.get(pid)
            if c is None:
                continue
            p = float(r[pcol])
            side, ev = O.pick_side(p, c["best_over"], c["best_under"])
            if side is None or ev < ev_threshold:
                continue
            cg = _as_int(r.get("CareerG"))
            note = ("rookie <50 G" if cg is not None and cg < 50
                    else ("1 book" if c["n_books"] == 1 else ""))
            match = {"ID": pid}
            if b_game:
                match["Game"] = r["Game"]
            add(_bet_rec(r["Game"] if b_game else single_tag, r["Name"],
                         r.get("Team", ""), meta["label"], side, meta["line"],
                         p, c, ev, note),
                "Batter Props", match, disp, side)

    # ---- pitcher count-prop lines (join on PlayerId, multi-line) ----
    starters = out["starters"]
    s_game = "Game" in starters.columns
    for skey, prefix in STARTER_PREFIX.items():
        meta = O.STARTER_MARKET.get(skey)
        if meta is None:
            continue
        for col in [c for c in starters.columns if c.startswith(prefix)]:
            try:
                line = float(col[len(prefix):])
            except ValueError:
                continue
            c_by_pid = cons(meta["api"], line, "PlayerId")
            if not c_by_pid:
                continue
            disp = _over_display(col)
            label = f'{meta["label"]} o{line:g}'
            for _, r in starters.iterrows():
                pid = _as_int(r.get("PlayerId"))
                c = c_by_pid.get(pid)
                if c is None:
                    continue
                p = float(r[col])
                side, ev = O.pick_side(p, c["best_over"], c["best_under"])
                if side is None or ev < ev_threshold:
                    continue
                match = {"ID": pid}
                if s_game:
                    match["Game"] = r["Game"]
                add(_bet_rec(r["Game"] if s_game else single_tag, r["Name"],
                             r.get("Team", ""), label, side, line, p, c, ev,
                             "1 book" if c["n_books"] == 1 else ""),
                    "Pitching Props", match, disp, side)

    # ---- game markets: totals + moneyline (join on home Team) ----
    games_sheet = "games" in out
    for g in _game_records(out, specs):
        for lstr, pv in g["totals"].items():
            line = float(lstr)
            c = cons("totals", line, "Team").get(g["home"])
            if c is None:
                continue
            p = float(pv)
            side, ev = O.pick_side(p, c["best_over"], c["best_under"])
            if side is None or ev < ev_threshold:
                continue
            add(_bet_rec(g["tag"], "Game total", g["home"], "total runs", side,
                         line, p, c, ev,
                         "1 book" if c["n_books"] == 1 else ""),
                "Games", {"Game": g["tag"]},
                (f"Runs > {lstr}" if games_sheet else None), side)
        c = cons("h2h", None, "Team").get(g["home"])
        if c is not None:
            p = g["home_win_prob"]
            side, ev = O.pick_side(p, c["best_over"], c["best_under"])
            if side is not None and ev >= ev_threshold:
                team = g["home"] if side == "over" else g["away"]
                add(_bet_rec(g["tag"], "Moneyline", "", "moneyline", side, None,
                             p, c, ev, "winner: no proven edge vs. always-home",
                             side_label=team),
                    "Games", {"Game": g["tag"]},
                    # Win Prob shows the HOME win %, so it is the "over" cell:
                    # paint it only when home is the pick; away bets live on
                    # the Bets sheet only.
                    ("Win Prob" if games_sheet else None), side)

    highlights = dict(highlights)
    if not bets:
        return (pd.DataFrame(columns=BET_HEADERS), highlights,
                f"Odds present for {', '.join(sorted(dates))}, but no side "
                f"cleared the EV >= {ev_threshold:.0%} bar. The model sees "
                f"no bet worth taking at these prices.")
    df = (pd.DataFrame(bets, columns=BET_HEADERS)
          .sort_values("EV%", ascending=False).reset_index(drop=True))
    return df, highlights, None


def _bets_sheet(bets_df, note):
    """The frame written to the 'Bets' sheet: the bet board, or a one-cell
    note when there are no bets / no odds."""
    if len(bets_df):
        return bets_df
    return pd.DataFrame({"Bets": [note or "No bets to show."]})


# ---------------------------------------------------------------------------
# Quality marks: light-blue highlights on the picks each column's own
# held-out diagnostics say are its best use — the prop-rankings playbook
# applied to today's slate. Data-driven, no odds for SELECTION.
#
# HOW DEEP a column is painted is decided in prop_rankings.quality_playbook()
# (single source of truth with the rankings workbook, so the two can never
# disagree) and is CI-AWARE end to end: a column earns depth only when the
# LOWER BOUND of its odds-ratio top-pick lift clears the floor and its
# calibration slope is sane, and that depth is then capped by the tier of its
# Score LOWER BOUND — the day-block-bootstrap Score_lo, not a point estimate.
# So a thin market (SB, a deep K line) can no longer buy deep blue on a lucky
# season: its CI is wide, its Score_lo falls, its depth shrinks. Odds-ratio
# lift (top-pick odds / base odds) rather than raw lift, because raw lift caps
# at 1/base — a 61%-base column like Hit could never clear a uniform raw bar
# no matter how real its selection power is.
#
# WHICH cells get painted is decided here, on today's slate:
#   - OVER side only, ranked by P(over) — the side the lift diagnostics
#     validate. (Ranking "confidence" via max(p, 1-p) used to mark trivial
#     unders on tail lines, e.g. a 1% K > 8.5 cell.) This also keeps blue
#     consistent with green, which is likewise over-only on the grids.
#   - each mark clears an INFORMEDNESS FLOOR over the column's own base rate,
#     so a high-base line is never marked just for being likely: K > 3.5
#     (base ~67%) needs P(over) >= 77.5%, K > 8.5 (base ~8%) needs >= 15.4%.
#   - where the odds store prices the exact market/line, the sharp consensus
#     is a VETO only (a pick the sharp books price under half the model's
#     number is dropped). Odds never SELECT a blue mark — that stays green.
#
# Diagnostics come LIVE from the paired snapshots + artifacts via
# prop_rankings (the same numbers PROP_RANKINGS prints, but never read from
# that xlsx — it goes stale; the snapshots refresh with every baseline). The
# bootstrap behind Score_lo is disk-cached by snapshot fingerprint, so the
# first call after a baseline pays ~1 min and every later serve is instant.
QUALITY_BASE_MULT, QUALITY_BASE_PAD = 2.0, 0.10
# where the odds store prices the exact market/line, the sharp consensus
# is a sanity anchor: a pick the sharp books price at less than half the
# model's number is dropped. Veto only — value SELECTION stays green/+EV.
QUALITY_FAIR_VETO = 0.5


def quality_marks(out, specs=None):
    """{sheet: [(row_match, display_column), ...]} of rank-quality picks to
    paint light blue. Depth per column comes from
    prop_rankings.quality_playbook() — CI-aware (odds-ratio lift LOWER bound
    + Score_lo tier cap), shared with the rankings workbook. This function
    only applies it to TODAY's slate: OVER side, top-`depth` by P(over), each
    clearing the informedness floor over that column's own base rate and — if
    the odds store prices the exact market/line — the sharp-consensus veto.
    Fails soft: any problem returns no marks (or just no veto), never blocks
    the workbook."""
    try:
        # prop_rankings lives in Tools/ (the manually-run toolkit); when
        # this runs under the GUI the Tools dir is already importable, but
        # a direct `python Model/predict.py` run needs it on the path.
        tools_dir = str(Path(__file__).resolve().parents[1] / "Tools")
        if tools_dir not in sys.path:
            sys.path.insert(0, tools_dir)
        import prop_rankings as R
        book = R.quality_playbook()
    except Exception:
        return {}
    # sharp fair probs for the slate, if odds were scraped (else no veto)
    store, cons_cache = None, {}
    try:
        s = O.load_odds(O.DEFAULT_STORE)
        if specs:
            dates = {str(sp["date"]) for sp in specs}
            s = s[s["Date"].astype(str).isin(dates)]
        if not s.empty:
            store = s
    except Exception:
        pass

    def sharp_fair(api, line, pid):
        if store is None or api is None or pid is None:
            return np.nan
        ck = (api, line)
        if ck not in cons_cache:
            try:
                cons_cache[ck] = _consensus(store, api, line, "PlayerId")
            except Exception:
                cons_cache[ck] = {}
        c = cons_cache[ck].get(pid)
        return c["fair"] if c else np.nan

    marks = defaultdict(list)

    def collect(frame, pcol, n, base, api, line, sheet, disp, has_game):
        """Top-n rows by P(over), each clearing the informedness floor and
        the sharp-line veto; a vetoed row frees its slot for the next."""
        floor = (min(QUALITY_BASE_MULT * base, base + QUALITY_BASE_PAD)
                 if np.isfinite(base) and 0 < base < 1 else 0.0)
        got = 0
        for _, r in frame.sort_values(pcol, ascending=False).iterrows():
            if got >= n:
                break
            p = float(r[pcol])
            if not np.isfinite(p) or p < floor:
                break                        # sorted desc — nothing left
            pid = _as_int(r.get("PlayerId"))
            fair = sharp_fair(api, line, pid)
            if np.isfinite(fair) and fair < QUALITY_FAIR_VETO * p:
                continue                     # fights the sharp line 2:1
            match = {"ID": pid}
            if has_game:
                match["Game"] = r["Game"]
            marks[sheet].append((match, disp))
            got += 1

    # ---- batter binaries: top-N of today's slate by probability ----
    batters = out["batters"]
    b_game = "Game" in batters.columns
    for prop, pcol in PROP_COLS.items():
        pb = book.get("binary", {}).get(prop)
        if pb is None or not pb["depth"] or pcol not in batters.columns:
            continue
        meta = O.PROP_MARKET.get(prop)
        collect(batters, pcol, pb["depth"], pb["base"],
                meta["api"] if meta else None,
                meta["line"] if meta else None,
                "Batter Props", BAT_HEADERS.get(pcol, pcol), b_game)

    # ---- pitcher O/U lines: per-line depth from the playbook ----
    starters = out["starters"]
    s_game = "Game" in starters.columns
    for skey, prefix in STARTER_PREFIX.items():
        ckey = R.QUAL_STARTER_KEY.get(skey)
        if ckey is None:
            continue
        meta = O.STARTER_MARKET.get(skey)
        for col in [c for c in starters.columns if c.startswith(prefix)]:
            try:
                line = float(col[len(prefix):])
            except ValueError:
                continue
            pb = book.get("pitch_line", {}).get((ckey, line))
            if pb is None or not pb["depth"]:
                continue
            collect(starters, col, pb["depth"], pb["base"],
                    meta["api"] if meta else None, line,
                    "Pitching Props", _over_display(col), s_game)
    return dict(marks)


def _cell_eq(cv, mv):
    """Match a written cell value against a highlight key (str Game tag or
    numeric PlayerId)."""
    if isinstance(mv, str):
        return str(cv) == mv
    ci = _as_int(cv)
    return ci is not None and ci == _as_int(mv)


def _polish(path, highlights=None, quality=None):
    """Make the workbook readable, in MLB colors (navy #041E42, red
    #BF0D3E): frozen bold headers, percent formats, autofit column widths
    (padded for the filter dropdown arrows), thin borders, white data
    cells (no column tints), a red flag on
    sub-50-game batters, and filter/sort dropdowns on the data sheets. The
    'Bets' board gets a green header + green rows, and `highlights` (from
    compute_bets) paints each flagged pick green on its prop sheet.
    `quality` (from quality_marks) paints rank-quality picks light blue —
    painted first, so a pick that is also +EV keeps the stronger green."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    wb = openpyxl.load_workbook(path)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="041E42")     # MLB navy
    bet_head_fill = PatternFill("solid", fgColor="1E7B34")  # Bets: green
    bet_row_fill = PatternFill("solid", fgColor="E7F3E2")   # light green board
    green_fill = PatternFill("solid", fgColor="92D050")     # flagged bet cell
    blue_fill = PatternFill("solid", fgColor="00B0F0")      # quality pick cell
    purple_fill = PatternFill("solid", fgColor="B1A0C7")    # quality AND +EV
    dim_font = Font(color="808B99")                        # ID column
    warn_font = Font(bold=True, color="BF0D3E")            # Career G < 50
    thin = Side(style="thin", color="B7C2D4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    # left-aligned text columns; everything else centers
    text_cols = {"Team", "Name", "Game", "Venue", "Winner",
                 "Predicted winner", "Expected score", "What",
                 "The model says", "Field", "Value", "Term", "Meaning"}

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        is_bets = ws.title == "Bets"
        filtered = ws.title in ("Batter Props", "Pitching Props", "Games") or \
            (is_bets and "EV%" in [str(c.value) for c in ws[1]])
        if filtered:
            ws.auto_filter.ref = ws.dimensions  # sort/filter dropdowns
        max_row, max_col = ws.max_row, ws.max_column

        headers = [str(c.value) for c in ws[1]]
        for j, cell in enumerate(ws[1], start=1):
            cell.font = head_font
            cell.fill = bet_head_fill if is_bets else head_fill
            cell.alignment = center if headers[j - 1] not in text_cols else left
            cell.border = border

        for j, h in enumerate(headers, start=1):
            is_text = h in text_cols or (is_bets and h in BET_TEXT_COLS)
            is_pct = " > " in h or h in PCT_COLS or (is_bets and h in BET_PCT_COLS)
            for i in range(2, max_row + 1):
                c = ws.cell(row=i, column=j)
                c.border = border
                c.alignment = left if is_text else center
                if is_pct:
                    c.number_format = "0.0%"
                    if isinstance(c.value, (int, float)) and c.value > 0.5:
                        c.font = Font(bold=True)
                if is_bets and h == "Best Odds":
                    c.number_format = "+0;-0"   # +475 / -150
                if h == "ID":
                    c.font = dim_font
                if is_bets:
                    c.fill = bet_row_fill        # the whole board reads "green"
                # sub-50-game batters are the least reliable picks (Glossary)
                if (h == "Career G" and isinstance(c.value, (int, float))
                        and c.value < 50):
                    c.font = warn_font

        # paint flagged picks. Resolve every mark to its concrete cell
        # first so overlaps combine: quality-only = blue, +EV-only = green,
        # BOTH (a rank-quality pick that is also a priced edge) = light
        # purple — the strongest signal on the sheet.
        cell_kind = {}
        for marks, kind in ((quality, "blue"), (highlights, "green")):
            for match, colname in (marks or {}).get(ws.title, []):
                hidx = {h: k + 1 for k, h in enumerate(headers)}
                if colname not in hidx or any(m not in hidx for m in match):
                    continue
                for i in range(2, max_row + 1):
                    if all(_cell_eq(ws.cell(row=i, column=hidx[m]).value, v)
                           for m, v in match.items()):
                        key = (i, hidx[colname])
                        prev = cell_kind.get(key)
                        cell_kind[key] = ("purple" if prev and prev != kind
                                          else kind)
                        break
        # pick fills keep their tinted text color, but probabilities inside
        # them follow the same bold-over-50% rule as plain cells; text
        # picks (e.g. Winner) stay bold for contrast on the fill
        paint = {"blue": (blue_fill, "0B2E4F"),
                 "green": (green_fill, "1E4620"),
                 "purple": (purple_fill, "3B2151")}
        for (i, j), kind in cell_kind.items():
            cc = ws.cell(row=i, column=j)
            fill, color = paint[kind]
            bold = not isinstance(cc.value, (int, float)) or cc.value > 0.5
            cc.fill, cc.font = fill, Font(bold=bold, color=color)

        # autofit: widest of header / any cell, clamped; percents count as
        # ~6 chars. The filter dropdown arrow is ~3 chars wide on the RIGHT
        # of the header cell, and headers are centered — so the header needs
        # arrow-width padding on BOTH sides to stay clear of it.
        pad = 6 if filtered else 0
        # the data boards stay compact (cap 90); the text reference sheets are
        # widened to fit their full content, up to Excel's 255-char column max
        cap = 255 if ws.title in ("Summary", "Glossary") else 90
        for j, h in enumerate(headers, start=1):
            longest = len(h) + pad
            for i in range(2, max_row + 1):
                v = ws.cell(row=i, column=j).value
                if v is None:
                    continue
                longest = max(longest, 6 if isinstance(v, float) else len(str(v)))
            ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = \
                min(max(longest + 2, 6), cap)
    wb.save(path)


def default_excel_path(spec):
    PRED_DIR.mkdir(exist_ok=True)
    import time as _t
    stamp = _t.strftime("%H%M%S")
    name = f'{spec["date"]}_{spec["away_team"]}_at_{spec["home_team"]}_{stamp}.xlsx'
    return PRED_DIR / name


def save_excel(spec, out, path=None):
    """Write one workbook per prediction run: game info, HR board, starters."""
    path = Path(path) if path else default_excel_path(spec)
    t = out["totals"]
    p_home = t.get("home_win_prob", 0.5)
    fav = spec["home_team"] if p_home >= 0.5 else spec["away_team"]
    info_rows = [
        ("Matchup", f'{spec["away_team"]} @ {spec["home_team"]}'),
        ("Date", spec["date"]), ("Stadium", spec["venue"]),
        ("Day/Night", spec["day_night"]), ("Temp (F)", spec["temp"]),
        ("Wind (mph)", spec["wind_speed"]), ("Wind dir", spec["wind_dir"]),
        ("Condition", spec["condition"]),
        ("Away starter", spec["away_starter"]),
        ("Home starter", spec["home_starter"]),
        ("Expected winner", f'{fav} {max(p_home, 1 - p_home):.0%} '
         f'(calibrated estimate, not a pick)'),
        ("Expected score",
         f'{spec["away_team"]} {t.get("exp_away_runs", "?")} — '
         f'{spec["home_team"]} {t.get("exp_home_runs", "?")}'),
        ("Expected lineup HRs", t["exp_lineup_HR"]),
        ("Expected total runs", t["exp_total_runs"]),
    ] + [(f"P(runs over {k})", v) for k, v in t["P_over_runs"].items()]
    info = pd.DataFrame(info_rows, columns=["Field", "Value"])
    summary = summary_frame([spec], out)
    batters = _display(out["batters"].sort_values("P_HR", ascending=False),
                       BAT_HEADERS, BAT_ORDER, drop=("xHR", "HR_fair_odds"))
    starters = _display(out["starters"].sort_values("xK", ascending=False),
                        ST_HEADERS, ST_ORDER)
    bets_df, highlights, note = compute_bets(out, [spec])
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        batters.to_excel(xw, sheet_name="Batter Props", index=False)
        starters.to_excel(xw, sheet_name="Pitching Props", index=False)
        info.to_excel(xw, sheet_name="Game", index=False)
        _bets_sheet(bets_df, note).to_excel(xw, sheet_name="Bets", index=False)
        summary.to_excel(xw, sheet_name="Summary", index=False)
        pd.DataFrame(GLOSSARY, columns=["Term", "Meaning"]).to_excel(
            xw, sheet_name="Glossary", index=False)
    _polish(path, highlights, quality=quality_marks(out, [spec]))
    return path


def _single_from_slate(out, i, tag):
    b, s = out["batters"], out["starters"]
    return {"batters": b[b["Game"] == tag].drop(columns=["Game"])
            .reset_index(drop=True),
            "starters": s[s["Game"] == tag].drop(columns=["Game"])
            .reset_index(drop=True),
            "totals": _slate_totals(out, i)}


def save_excel_slate(specs, out, path=None):
    """Combined slate workbook: Summary, per-game totals, cross-game boards."""
    if len(specs) == 1:
        return save_excel(specs[0], _single_from_slate(
            out, 0, f'{specs[0]["away_team"]}@{specs[0]["home_team"]}'), path)
    if path is None:
        PRED_DIR.mkdir(exist_ok=True)
        import time as _t
        date = min(s["date"] for s in specs)
        path = PRED_DIR / (f'{date}_slate_{len(specs)}games_'
                           f'{_t.strftime("%H%M%S")}.xlsx')
    # every sheet in slate order (the GUI slate = earliest first pitch
    # first, or however the user arranged it with the move buttons), then
    # the sheet's own metric within each game
    order = {f'{s["away_team"]}@{s["home_team"]}': i
             for i, s in enumerate(specs)}

    def by_game(df, metric):
        go = df["Game"].map(order).fillna(len(order))
        return (df.assign(_go=go)
                .sort_values(["_go", metric], ascending=[True, False])
                .drop(columns="_go"))

    batters = _display(by_game(out["batters"], "P_HR"),
                       BAT_HEADERS, BAT_ORDER, drop=("xHR", "HR_fair_odds"))
    starters = _display(by_game(out["starters"], "xK"),
                        ST_HEADERS, ST_ORDER)
    games = _display(by_game(out["games"], "WinProb"),
                     GAME_HEADERS, GAME_ORDER, drop=("HomeWinProb",))
    bets_df, highlights, note = compute_bets(out, specs)
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        batters.to_excel(xw, sheet_name="Batter Props", index=False)
        starters.to_excel(xw, sheet_name="Pitching Props", index=False)
        games.to_excel(xw, sheet_name="Games", index=False)
        _bets_sheet(bets_df, note).to_excel(xw, sheet_name="Bets", index=False)
        summary_frame(specs, out).to_excel(xw, sheet_name="Summary", index=False)
        pd.DataFrame(GLOSSARY, columns=["Term", "Meaning"]).to_excel(
            xw, sheet_name="Glossary", index=False)
    _polish(path, highlights, quality=quality_marks(out, specs))
    return Path(path)


def save_excel_per_game(specs, out, folder=None):
    """One workbook per game of a slate. Returns the list of paths."""
    if folder is None:
        PRED_DIR.mkdir(exist_ok=True)
        folder = PRED_DIR
    folder = Path(folder)
    folder.mkdir(exist_ok=True)
    paths = []
    for i, spec in enumerate(specs):
        tag = f'{spec["away_team"]}@{spec["home_team"]}'
        p = folder / (f'{spec["date"]}_{spec["away_team"]}_at_'
                      f'{spec["home_team"]}.xlsx')
        n = 2
        while p in paths or (p.exists() and p not in paths):
            p = folder / (f'{spec["date"]}_{spec["away_team"]}_at_'
                          f'{spec["home_team"]}_{n}.xlsx')
            n += 1
        save_excel(spec, _single_from_slate(out, i, tag), p)
        paths.append(p)
    return paths


def _slate_totals(out, i):
    g = out["games"].iloc[i]
    return {"exp_lineup_HR": g["exp_lineup_HR"],
            "exp_total_runs": g["exp_total_runs"],
            "exp_away_runs": g["exp_away_runs"],
            "exp_home_runs": g["exp_home_runs"],
            "home_win_prob": g["HomeWinProb"],
            "P_over_runs": {k.replace("P_runs_over_", ""): g[k]
                            for k in out["games"].columns
                            if k.startswith("P_runs_over_")},
            "P_over_away_runs": {k.replace("P_away_runs_over_", ""): g[k]
                                 for k in out["games"].columns
                                 if k.startswith("P_away_runs_over_")},
            "P_over_home_runs": {k.replace("P_home_runs_over_", ""): g[k]
                                 for k in out["games"].columns
                                 if k.startswith("P_home_runs_over_")}}


# ------------------------------------------------------------ replay/test


def spec_from_game(stores, gamepk):
    """Rebuild the input spec of a real game from the data (for replay/tests)."""
    r = stores.raw
    g = r["games"][r["games"]["GamePk"] == gamepk].iloc[0]
    gb = r["gb"][r["gb"]["GamePk"] == gamepk]
    gp = r["gp"][(r["gp"]["GamePk"] == gamepk) & (r["gp"]["GS"] == 1)]

    def lineup(team):
        rows = gb[(gb["Team"] == team) & gb["BattingOrder"].notna()].copy()
        rows["bo"] = pd.to_numeric(rows["BattingOrder"], errors="coerce")
        rows = rows[rows["bo"] % 100 == 0].sort_values("bo")
        return [(int(p), int(b // 100)) for p, b in zip(rows["PlayerId"], rows["bo"])]

    def starter(team):
        m = gp[gp["Team"] == team]
        return int(m["PlayerId"].iloc[0]) if len(m) else None

    hp_ump = None                     # real HP ump, so replay/selftest match
    um = r.get("umps")
    if um is not None:
        mu = um[um["GamePk"] == gamepk]
        if len(mu) and pd.notna(mu["HpUmpId"].iloc[0]):
            hp_ump = int(mu["HpUmpId"].iloc[0])

    humidity = pressure = precip = None   # real weather, same reason
    wt = r.get("weather")
    if wt is not None:
        mw = wt[wt["GamePk"] == gamepk]
        if len(mw):
            humidity = mw["Humidity"].iloc[0]
            pressure = mw["Pressure"].iloc[0]
            precip = mw["Precip"].iloc[0]

    # doubleheader flags, derived exactly like the training table (#24):
    # games per (Date, HomeTeam), game 2 = the higher GamePk of the day
    day = r["games"][(r["games"]["Date"] == g["Date"])
                     & (r["games"]["HomeTeam"] == g["HomeTeam"])]
    pks = sorted(day["GamePk"].tolist())
    is_dh = float(len(pks) >= 2)
    dh_game2 = float(pks.index(gamepk) >= 1) if gamepk in pks else 0.0

    return {
        "date": str(g["Date"].date()), "away_team": g["AwayTeam"],
        "home_team": g["HomeTeam"], "venue": g["Venue"],
        "day_night": g["DayNight"], "temp": g["Temp"],
        "wind_speed": g["WindSpeed"], "wind_dir": g["WindDir"],
        "condition": g["Condition"],
        "humidity": humidity, "pressure": pressure, "precip": precip,
        "away_starter": starter(g["AwayTeam"]),
        "home_starter": starter(g["HomeTeam"]),
        "away_lineup": lineup(g["AwayTeam"]),
        "home_lineup": lineup(g["HomeTeam"]),
        "hp_ump_id": hp_ump,
        "is_dh": is_dh, "dh_game2": dh_game2,
    }


def _compare_row(train_row, serve_row, cols, label, worst, stats):
    """Fold one train-vs-serve row comparison into (worst, stats).
    stats counts checked / both-NaN (unverifiable — audit #12: these used
    to vanish silently, hiding features broken to NaN in both paths) /
    NaN-mismatch pairs."""
    for c in cols:
        a = train_row.get(c, np.nan)
        b = serve_row.get(c, np.nan)
        a = np.nan if pd.isna(a) else float(a)
        b = np.nan if pd.isna(b) else float(b)
        if np.isnan(a) and np.isnan(b):
            stats["both_nan"] += 1
            continue
        if np.isnan(a) != np.isnan(b):
            stats["nan_mismatch"] += 1
            print(f"  NaN mismatch {label} {c}: train={a} serve={b}")
            continue
        d = abs(a - b) / max(1e-9, abs(a))
        stats["checked"] += 1
        if d > worst[1]:
            worst = (f"{label} {c} (train={a:.6g} serve={b:.6g})", d)
    return worst, stats


def _selftest_games(bf):
    """GamePks to parity-check (audit #12): a one-game spot check passes
    trivially for features broken only in contexts that game lacks. Cover
    the newest game PLUS the newest doubleheader game 2 and the newest
    renamed-franchise (ATH) game when the frames have them."""
    season = int(bf["Season"].max())
    recent = bf[(bf["Season"] == season) & bf["StarterId"].notna()]
    picks = [int(recent["GamePk"].iloc[-1])]
    if "dh_game2" in recent.columns:
        dh = recent[recent["dh_game2"] == 1.0]
        if len(dh):
            picks.append(int(dh["GamePk"].iloc[-1]))
    ath = recent[recent["Team"] == "ATH"]
    if len(ath):
        picks.append(int(ath["GamePk"].iloc[-1]))
    return list(dict.fromkeys(picks))


def selftest(pred):
    """Compare inference-path features against the training frames for real
    games from the newest season: they must match, or training and serving
    have drifted. Covers batter rows, starter (K-model) rows, the winner
    row, and the team-game rows, across several game CONTEXTS (plain, DH
    game 2, renamed team), and reports how many pairs were unverifiable
    (NaN in both paths) instead of silently skipping them."""
    frames = joblib.load(ART / "frames.joblib")
    bf = frames["bf"]
    worst = ("", 0.0)
    stats = {"checked": 0, "both_nan": 0, "nan_mismatch": 0}
    check_cols = [c for c in F.batter_feature_cols() if c not in F.CAT_COLS]

    for gamepk in _selftest_games(bf):
        spec = spec_from_game(pred.stores, gamepk)
        print(f"selftest on GamePk {gamepk}: {spec['away_team']} @ "
              f"{spec['home_team']} {spec['date']}"
              + (" [DH game 2]" if spec.get("dh_game2") else ""))

        bdf, bmeta = pred._batter_rows(spec)
        # the FULL batter superset (not just the shipped model's selected
        # subset), so every served feature is parity-verified
        for i, m in bmeta.iterrows():
            trow = bf[(bf["GamePk"] == gamepk)
                      & (bf["PlayerId"] == m["PlayerId"])]
            if trow.empty:
                continue
            worst, stats = _compare_row(trow.iloc[0], bdf.iloc[i],
                                        check_cols, m["Name"], worst, stats)

        if "sf" in frames:
            sf = frames["sf"]
            sdf, smeta = pred._starter_rows(spec, bdf, bmeta)
            st_cols = [c for c in F.starts_feature_cols()
                       if c not in F.CAT_COLS and c in sf.columns]
            for i, m in smeta.iterrows():
                trow = sf[(sf["GamePk"] == gamepk)
                          & (sf["PlayerId"] == m["PlayerId"])]
                if trow.empty:
                    continue
                worst, stats = _compare_row(trow.iloc[0], sdf.iloc[i],
                                            st_cols, f'K:{m["Name"]}',
                                            worst, stats)

        if "win_model" in pred.art and "gf" in frames:
            grow = frames["gf"][frames["gf"]["GamePk"] == gamepk]
            if not grow.empty:
                wcols = [c for c in F.win_feature_cols()
                         if c not in F.CAT_COLS and c in grow.columns]
                worst, stats = _compare_row(
                    grow.iloc[0], pred._win_row(spec, bdf, bmeta),
                    wcols, "win-row", worst, stats)

        # team-game rows: the totals surface's lineup/sched/luck columns
        if "gf" in frames:
            grow = frames["gf"][frames["gf"]["GamePk"] == gamepk]
            if not grow.empty:
                tg = F.build_team_game_frame(grow)
                gdf = pred._team_rows(spec, bdf, bmeta)
                tg_cols = [c for c in F.team_game_feature_cols()
                           if c not in F.CAT_COLS and c in tg.columns]
                for i, home in ((0, 0), (1, 1)):
                    trow = tg[tg["Home"] == home]
                    if trow.empty:
                        continue
                    worst, stats = _compare_row(
                        trow.iloc[0], gdf.iloc[i], tg_cols,
                        f"team-row[{'home' if home else 'away'}]",
                        worst, stats)

    tot = stats["checked"] + stats["both_nan"]
    print(f"  compared {stats['checked']} feature values "
          f"({stats['both_nan']} pairs NaN in both paths — unverifiable, "
          f"{stats['both_nan'] / max(tot, 1):.0%} of the surface; "
          f"{stats['nan_mismatch']} NaN mismatches); "
          f"worst relative diff: {worst[1]:.2e} [{worst[0]}]")
    ok = worst[1] < 1e-6 and stats["nan_mismatch"] == 0
    print("  PARITY OK" if ok else "  PARITY FAILED")
    return ok


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--game", type=int, help="replay a historical GamePk")
    ap.add_argument("--selftest", action="store_true")
    ap.add_argument("--recal", action="store_true",
                    help="apply stored in-season drift offsets to batter props")
    args = ap.parse_args()

    pred = Predictor(recal=args.recal)
    if args.selftest:
        sys.exit(0 if selftest(pred) else 1)

    if args.game:
        spec = spec_from_game(pred.stores, args.game)
        out = pred.predict_game(spec)
        print(f"\n{spec['away_team']} @ {spec['home_team']}  {spec['date']}  "
              f"{spec['venue']}  {spec['temp']}F wind {spec['wind_speed']} "
              f"{spec['wind_dir']}\n")
        print("=== Home run probabilities (lineup, ranked) ===")
        print(out["batters"].to_string(index=False))
        print("\n=== Starter strikeouts ===")
        print(out["starters"].to_string(index=False))
        print("\n=== Game totals ===")
        for k, v in out["totals"].items():
            print(f"  {k}: {v}")
        path = save_excel(spec, out)
        print(f"\nsaved workbook: {path}")


if __name__ == "__main__":
    main()
